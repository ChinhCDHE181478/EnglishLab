# -*- coding: utf-8 -*-
"""Rename IT_TCHHW sheet/IDs -> IT_MNGHW; refresh Test Cases + Statistics."""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
import rebuild_it_excel_from_uc as reb  # noqa: E402

EXCEL = Path(
    r"C:\Users\phong\Downloads\intergration test"
    r"\SEP490_G23_Report5.2_Integration Test_COMPLETED.xlsx"
)
PROJ = Path(r"D:\EngLishLab\EnglishLab\outputs\integration-test")
BACKUP = EXCEL.with_name(EXCEL.stem + "_BEFORE_RENAME_MNGHW.xlsx")

ID_MAP = {
    "IT_TCHHW_01": "IT_MNGHW_01",
    "IT_TCHHW_02": "IT_MNGHW_02",
}


def main():
    if not EXCEL.exists():
        raise SystemExit(f"Missing {EXCEL}")
    shutil.copy2(EXCEL, BACKUP)
    wb = load_workbook(EXCEL)

    if "IT_TCHHW" in wb.sheetnames:
        if "IT_MNGHW" in wb.sheetnames:
            wb.remove(wb["IT_MNGHW"])
        wb["IT_TCHHW"].title = "IT_MNGHW"
        print("renamed sheet IT_TCHHW -> IT_MNGHW")
    elif "IT_MNGHW" not in wb.sheetnames:
        raise SystemExit("Neither IT_TCHHW nor IT_MNGHW found")

    ws = wb["IT_MNGHW"]
    for row in range(11, min(ws.max_row or 11, 100) + 1):
        v = ws.cell(row, 1).value
        if isinstance(v, str) and v in ID_MAP:
            ws.cell(row, 1).value = ID_MAP[v]
            print(f"A{row}: {v} -> {ID_MAP[v]}")

    # Fix any leftover hyperlinks/text in Test Cases pointing to IT_TCHHW
    if "Test Cases" in wb.sheetnames:
        tc = wb["Test Cases"]
        for row in range(1, min(tc.max_row or 1, 80) + 1):
            for col in range(1, 8):
                cell = tc.cell(row, col)
                if isinstance(cell.value, str) and "IT_TCHHW" in cell.value:
                    cell.value = cell.value.replace("IT_TCHHW", "IT_MNGHW")
                if cell.hyperlink and getattr(cell.hyperlink, "location", None):
                    loc = cell.hyperlink.location or ""
                    if "IT_TCHHW" in loc:
                        cell.hyperlink.location = loc.replace("IT_TCHHW", "IT_MNGHW")
                if cell.hyperlink and getattr(cell.hyperlink, "display", None):
                    if cell.hyperlink.display == "IT_TCHHW":
                        cell.hyperlink.display = "IT_MNGHW"

    reb.rewrite_test_cases_sheet(wb["Test Cases"])
    reb.rewrite_statistics(wb["Test Statistics"])
    reb.reorder_it_sheets(wb)

    keep = {"Cover", "Test Cases", "Test Statistics"} | {
        m["sheet"] for m in reb.MODULES
    }
    for name in list(wb.sheetnames):
        if name not in keep:
            wb.remove(wb[name])
            print("REMOVED", name)

    try:
        wb.save(EXCEL)
        out = EXCEL
    except PermissionError:
        out = EXCEL.with_name(EXCEL.stem + "_MNGHW.xlsx")
        wb.save(out)
        print("LOCKED ->", out)
    shutil.copy2(out, PROJ / out.name)
    print("WROTE", out)


if __name__ == "__main__":
    main()
