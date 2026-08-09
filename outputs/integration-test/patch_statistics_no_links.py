# -*- coding: utf-8 -*-
"""Patch Test Statistics: remove sheet links; keep Pass/Fail formulas."""
from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

from full_modules import MODULES

ROOT = Path(r"C:\Users\phong\Downloads\intergration test")
PROJ = Path(r"D:\EngLishLab\EnglishLab\outputs\integration-test")
SRC = PROJ / "SEP490_G23_Report5.2_Integration Test_COMPLETED_v4.xlsx"
OUT = ROOT / "SEP490_G23_Report5.2_Integration Test_COMPLETED_v4.xlsx"
OUT_PROJ = PROJ / "SEP490_G23_Report5.2_Integration Test_COMPLETED_v4.xlsx"
VALUE_FONT = Font(name="Tahoma", size=10, color="008000")

wb = load_workbook(SRC)
st = wb["Test Statistics"]

# drop hyperlinks on module-code column
wanted = {f"C{r}" for r in range(11, 50)}
st._hyperlinks = [h for h in list(getattr(st, "_hyperlinks", []) or []) if (getattr(h, "ref", "") or "").upper() not in wanted]

for i, m in enumerate(MODULES, start=1):
    r = 10 + i
    sheet = m["sheet"]
    cell = st.cell(r, 3)
    cell.value = f"='{sheet}'!B2"
    cell.hyperlink = None
    cell.font = VALUE_FONT
    st.cell(r, 4).value = f"='{sheet}'!B6"
    st.cell(r, 5).value = f"='{sheet}'!C6"
    st.cell(r, 6).value = f"='{sheet}'!D6"
    st.cell(r, 7).value = f"='{sheet}'!E6"
    st.cell(r, 8).value = f"='{sheet}'!B4"

last = 10 + len(MODULES)
sub_r = last + 2
# clear any leftover links/rows between last module and subtotal area already set
for r in range(last + 1, sub_r):
    for c in range(2, 9):
        st.cell(r, c).value = None
        st.cell(r, c).hyperlink = None

st.cell(sub_r, 3).value = "Sub total"
st.cell(sub_r, 4).value = f"=SUM(D11:D{last})"
st.cell(sub_r, 5).value = f"=SUM(E11:E{last})"
st.cell(sub_r, 6).value = f"=SUM(F11:F{last})"
st.cell(sub_r, 7).value = f"=SUM(G11:G{last})"
st.cell(sub_r, 8).value = f"=SUM(H11:H{last})"

cov_r = sub_r + 2
st.cell(cov_r, 3).value = "Test coverage"
st.cell(cov_r, 5).value = f"=IF((H{sub_r}-G{sub_r})=0,0,(D{sub_r}+E{sub_r})*100/(H{sub_r}-G{sub_r}))"
st.cell(cov_r, 6).value = "%"
st.cell(cov_r + 1, 3).value = "Test successful coverage"
st.cell(cov_r + 1, 5).value = f"=IF((H{sub_r}-G{sub_r})=0,0,D{sub_r}*100/(H{sub_r}-G{sub_r}))"
st.cell(cov_r + 1, 6).value = "%"

wb.save(OUT)
shutil.copy2(OUT, OUT_PROJ)
print("patched", OUT)
print("module rows", last - 10, "subtotal", sub_r)
