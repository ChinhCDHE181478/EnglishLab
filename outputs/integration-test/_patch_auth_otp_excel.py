# -*- coding: utf-8 -*-
"""Set IT_AUTH_03 / IT_AUTH_09 to Passed after OTP-from-DB MockMvc IT."""
from __future__ import annotations

import shutil
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

EXCEL = Path(
    r"C:\Users\phong\Downloads\intergration test"
    r"\SEP490_G23_Report5.2_Integration Test_COMPLETED.xlsx"
)
PROJ = Path(r"D:\EngLishLab\EnglishLab\outputs\integration-test")
BACKUP = EXCEL.with_name(EXCEL.stem + "_BEFORE_AUTH_OTP_PASS.xlsx")

TODAY = date.today().isoformat()
TESTER = "phongdx"
NOTES = {
    "IT_AUTH_03": "Passed via MockMvc: register → read EMAIL_VERIFICATION OTP from auth_tokens → POST /api/auth/verify-email",
    "IT_AUTH_09": "Passed via MockMvc: register+verify → forgot → read PASSWORD_RESET OTP from auth_tokens → reset-password → login new password",
}

CASE_FILL = PatternFill("solid", fgColor="FFFFFF")
THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


def main():
    shutil.copy2(EXCEL, BACKUP)
    wb = load_workbook(EXCEL)
    ws = wb["IT_AUTH"]
    for row in range(11, min(ws.max_row or 11, 80) + 1):
        cid = ws.cell(row, 1).value
        if cid not in NOTES:
            continue
        for col_status, col_date, col_tester in ((6, 7, 8), (9, 10, 11), (12, 13, 14)):
            ws.cell(row, col_status).value = "Passed"
            ws.cell(row, col_date).value = TODAY
            ws.cell(row, col_tester).value = TESTER
        note = ws.cell(row, 15)
        note.value = NOTES[cid]
        note.fill = CASE_FILL
        note.font = Font(name="Tahoma", size=10)
        note.border = THIN
        note.alignment = Alignment(wrap_text=True, vertical="top")
        print("UPDATED", cid, "row", row)

    try:
        wb.save(EXCEL)
        out = EXCEL
    except PermissionError:
        out = EXCEL.with_name(EXCEL.stem + "_AUTH_OTP_PASS.xlsx")
        wb.save(out)
        print("LOCKED ->", out)
    shutil.copy2(out, PROJ / out.name)
    print("WROTE", out)


if __name__ == "__main__":
    main()
