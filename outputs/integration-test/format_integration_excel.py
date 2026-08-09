# -*- coding: utf-8 -*-
"""Format Integration Test COMPLETED_v2 workbook for readability (no content/structure changes)."""
from __future__ import annotations

import math
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(r"C:\Users\phong\Downloads\intergration test")
SRC = ROOT / "SEP490_G23_Report5.2_Integration Test_COMPLETED_HONEST.xlsx"
BACKUP = ROOT / "SEP490_G23_Report5.2_Integration Test_COMPLETED_HONEST_BACKUP.xlsx"
OUT = ROOT / "SEP490_G23_Report5.2_Integration Test_COMPLETED_HONEST_FORMATTED.xlsx"
OUT_PROJECT = Path(r"D:\EngLishLab\EnglishLab\outputs\integration-test") / OUT.name

THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
HEADER_FILL = PatternFill(start_color="76923C", end_color="76923C", fill_type="solid")
TC_HEADER_FILL = PatternFill(start_color="333399", end_color="333399", fill_type="solid")
STATS_HEADER_FILL = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
GROUP_FILL = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)
HEADER_FONT = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")
GROUP_FONT = Font(name="Tahoma", size=10, bold=True, color="000000")
BODY_FONT = Font(name="Tahoma", size=10, bold=False, color="000000")
LABEL_FONT = Font(name="Tahoma", size=10, bold=True, color="993300")
VALUE_FONT = Font(name="Tahoma", size=10, bold=False, color="008000")
LINK_FONT = Font(name="Tahoma", size=10, bold=False, color="0000FF", underline="single")
SUBTOTAL_FONT = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")

# Unified widths for IT module sheets (within requested ranges)
IT_WIDTHS = {
    "A": 16,   # ID
    "B": 40,   # Description
    "C": 55,   # Procedure
    "D": 50,   # Expected
    "E": 35,   # Pre-conditions
    "F": 12,   # Round 1
    "G": 13,   # date
    "H": 15,   # Tester
    "I": 12,
    "J": 13,
    "K": 15,
    "L": 12,
    "M": 13,
    "N": 15,
    "O": 25,   # Note
}

WRAP_TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)


def chars_per_line(width: float) -> int:
    return max(8, int(float(width) * 1.05))


def count_wrapped_lines(text: str | None, width: float) -> int:
    if text is None or str(text).strip() == "":
        return 1
    s = str(text)
    cpl = chars_per_line(width)
    total = 0
    for part in s.split("\n"):
        if part == "":
            total += 1
            continue
        total += max(1, math.ceil(len(part) / cpl))
    return max(1, total)


def estimate_row_height(ws, row: int, widths: dict) -> float:
    # Consider B,C,D,E,O primarily
    cols = [("B", widths["B"]), ("C", widths["C"]), ("D", widths["D"]), ("E", widths["E"]), ("O", widths["O"])]
    lines = 1
    for letter, w in cols:
        lines = max(lines, count_wrapped_lines(ws[f"{letter}{row}"].value, w))
    # also ID if long
    lines = max(lines, count_wrapped_lines(ws[f"A{row}"].value, widths["A"]))
    height = 14 + lines * 13.2
    return max(36.0, min(height, 260.0))


def last_table_row(ws, start=11) -> int:
    last = start - 1
    for r in range(start, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 16)]
        if any(v is not None and str(v).strip() != "" for v in vals):
            last = r
        elif last >= start and r > last + 5:
            break
    return last if last >= start else start


def apply_it_sheet(ws):
    for col, width in IT_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # Meta labels A2:A8 style lightly (no content change)
    for r in range(2, 9):
        a = ws.cell(r, 1)
        b = ws.cell(r, 2)
        if a.value is not None:
            a.font = LABEL_FONT
            a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if b.value is not None:
            b.font = BODY_FONT if r != 2 else BODY_FONT
            b.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if r == 3:
                b.alignment = WRAP_TOP_LEFT
                # requirement can be long
                lines = count_wrapped_lines(b.value, 60)
                ws.row_dimensions[r].height = max(22, min(14 + lines * 13, 90))

    # Header row 10
    ws.row_dimensions[10].height = 30
    for c in range(1, 16):
        cell = ws.cell(10, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_CENTER
        cell.border = THIN

    last = last_table_row(ws)
    for r in range(11, last + 1):
        a_val = ws.cell(r, 1).value
        is_group = isinstance(a_val, str) and a_val.strip() and not a_val.startswith("IT_")
        is_case = isinstance(a_val, str) and a_val.startswith("IT_")
        empty_row = a_val is None or str(a_val).strip() == ""

        for c in range(1, 16):
            cell = ws.cell(r, c)
            cell.border = THIN
            letter = get_column_letter(c)

            if is_group:
                cell.font = GROUP_FONT
                cell.fill = GROUP_FILL
                if c == 1:
                    cell.alignment = LEFT_CENTER
                else:
                    cell.alignment = LEFT_CENTER
            elif is_case:
                # Always white for test-case rows (do not keep leftover template cyan fills).
                cell.fill = WHITE_FILL
                cell.font = BODY_FONT
                if letter == "A":
                    cell.alignment = CENTER
                    cell.font = Font(name="Tahoma", size=10, bold=True, color="000000")
                elif letter in ("B", "C", "D", "E", "O"):
                    cell.alignment = WRAP_TOP_LEFT
                elif letter in ("F", "G", "H", "I", "J", "K", "L", "M", "N"):
                    cell.alignment = CENTER
                else:
                    cell.alignment = WRAP_TOP_LEFT
            elif not empty_row:
                cell.font = BODY_FONT
                cell.fill = WHITE_FILL
                if letter == "A":
                    cell.alignment = CENTER
                elif letter in ("B", "C", "D", "E", "O"):
                    cell.alignment = WRAP_TOP_LEFT
                elif letter in ("F", "G", "H", "I", "J", "K", "L", "M", "N"):
                    cell.alignment = CENTER
                else:
                    cell.alignment = WRAP_TOP_LEFT
            else:
                # Clear leftover fills on blank trailing rows inside bordered region
                cell.fill = NO_FILL
                cell.border = THIN

        if is_group:
            ws.row_dimensions[r].height = 24
        elif is_case:
            ws.row_dimensions[r].height = estimate_row_height(ws, r, IT_WIDTHS)
        else:
            ws.row_dimensions[r].height = 22

    return last


def format_test_cases(ws):
    # Columns B-F used in list
    widths = {"B": 8, "C": 32, "D": 18, "E": 55, "F": 42}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.column_dimensions["A"].width = 3

    # Header row 8 — template blue #333399
    for c in range(2, 7):
        cell = ws.cell(8, c)
        if cell.value is not None:
            cell.font = HEADER_FONT
            cell.fill = TC_HEADER_FILL
            cell.alignment = WRAP_CENTER
            cell.border = THIN
    ws.row_dimensions[8].height = 28

    # Title
    if ws["D1"].value:
        ws["D1"].font = Font(name="Tahoma", size=14, bold=True)
        ws["D1"].alignment = CENTER

    for r in range(3, 7):
        for c in range(2, 7):
            cell = ws.cell(r, c)
            if cell.value is not None:
                if c == 2:
                    cell.font = LABEL_FONT
                elif c == 4:
                    cell.font = VALUE_FONT
                else:
                    cell.font = BODY_FONT
                cell.alignment = WRAP_TOP_LEFT if c >= 4 and r == 5 else LEFT_CENTER
                if r == 5 and c == 4:
                    lines = count_wrapped_lines(cell.value, widths.get("E", 40))
                    ws.row_dimensions[r].height = max(40, min(14 + lines * 12, 120))

    # Data rows from 9
    last = 8
    for r in range(9, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(2, 7)):
            last = r
        elif last > 8 and r > last + 3:
            break
    for r in range(9, last + 1):
        for c in range(2, 7):
            cell = ws.cell(r, c)
            cell.border = THIN
            # Sheet Name column must stay blue underlined link
            if c == 4 and (cell.hyperlink or (isinstance(cell.value, str) and cell.value.startswith("IT"))):
                cell.font = LINK_FONT
            else:
                cell.font = BODY_FONT
            if c == 2:
                cell.alignment = CENTER
            elif c in (3, 4):
                cell.alignment = LEFT_CENTER
            else:
                cell.alignment = WRAP_TOP_LEFT
        lines = max(
            count_wrapped_lines(ws.cell(r, 5).value, widths["E"]),
            count_wrapped_lines(ws.cell(r, 6).value, widths["F"]),
            1,
        )
        ws.row_dimensions[r].height = max(28, min(14 + lines * 12, 140))


def format_statistics(ws):
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 18

    if ws["B1"].value:
        ws["B1"].font = Font(name="Tahoma", size=14, bold=True)
        ws["B1"].alignment = CENTER

    # Header row 10 — template navy #000080
    for c in range(2, 9):
        cell = ws.cell(10, c)
        cell.font = HEADER_FONT
        cell.fill = STATS_HEADER_FILL
        cell.alignment = WRAP_CENTER
        cell.border = THIN
    ws.row_dimensions[10].height = 28

    # Find last data/subtotal row
    last = 10
    for r in range(11, min(ws.max_row, 80) + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(2, 9)):
            last = r
    for r in range(11, last + 1):
        is_subtotal = str(ws.cell(r, 3).value or "").strip().lower() == "sub total"
        for c in range(2, 9):
            cell = ws.cell(r, c)
            cell.border = THIN
            if is_subtotal:
                cell.font = SUBTOTAL_FONT
                cell.fill = STATS_HEADER_FILL
                cell.alignment = CENTER if c != 3 else LEFT_CENTER
            else:
                cell.font = BODY_FONT
                cell.fill = WHITE_FILL
                cell.alignment = CENTER if c != 3 else LEFT_CENTER
                # Count columns must be numeric (not %), leftover template formats break module rows
                if c >= 4:
                    cell.number_format = "0"
                else:
                    cell.number_format = "General"
        ws.row_dimensions[r].height = 22

    # Meta rows — brown labels / green values like template
    label_values = {
        "Project Name", "Project Code", "Document Code", "Notes",
        "Creator", "Reviewer/Approver", "Issue Date",
        "Test coverage", "Test successful coverage",
    }
    for r in range(3, last + 5):
        for c in range(2, 9):
            cell = ws.cell(r, c)
            if cell.value is None:
                continue
            text = str(cell.value)
            if text in label_values:
                cell.font = LABEL_FONT
            elif c in (3, 4, 6, 8) and r <= 6 and not text.startswith("=") and text not in label_values:
                if c == 3 and r in (3, 4, 5, 6):
                    cell.font = VALUE_FONT
            if r in (3, 4, 5, 6) and c in (3, 6, 8) and text.startswith("="):
                cell.font = VALUE_FONT


def format_cover(ws):
    # Light touch: wrap and row heights for change log
    for col, w in [("A", 18), ("B", 22), ("C", 36), ("D", 10), ("E", 45), ("F", 40)]:
        ws.column_dimensions[col].width = w
    if ws["B2"].value:
        ws["B2"].font = Font(name="Tahoma", size=16, bold=True)
        ws["B2"].alignment = CENTER
    for r in range(4, 7):
        for c in range(1, 7):
            cell = ws.cell(r, c)
            if cell.value is not None:
                cell.font = LABEL_FONT if c in (1, 5) else BODY_FONT
                cell.alignment = LEFT_CENTER
    # Record of change header
    for c in range(1, 7):
        cell = ws.cell(10, c)
        if cell.value is not None:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = WRAP_CENTER
            cell.border = THIN
    ws.row_dimensions[10].height = 28
    for r in range(11, 20):
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, 7)):
            for c in range(1, 7):
                cell = ws.cell(r, c)
                cell.border = THIN
                cell.font = BODY_FONT
                cell.alignment = WRAP_TOP_LEFT if c in (3, 5, 6) else CENTER
            lines = max(count_wrapped_lines(ws.cell(r, 5).value, 45), count_wrapped_lines(ws.cell(r, 6).value, 40))
            ws.row_dimensions[r].height = max(28, min(14 + lines * 12, 100))


def main():
    shutil.copy2(SRC, BACKUP)
    shutil.copy2(SRC, OUT)
    wb = load_workbook(OUT)
    sheet_order = list(wb.sheetnames)

    it_sheets = []
    borders_sheets = []
    for name in sheet_order:
        ws = wb[name]
        if name.startswith("IT - "):
            last = apply_it_sheet(ws)
            it_sheets.append(name)
            borders_sheets.append(name)
        elif name == "Test Cases":
            format_test_cases(ws)
            borders_sheets.append(name)
        elif name == "Test Statistics":
            format_statistics(ws)
            borders_sheets.append(name)
        elif name == "Cover":
            format_cover(ws)
            borders_sheets.append(name)

    assert list(wb.sheetnames) == sheet_order

    # Formula / REF check
    refs = []
    for name in wb.sheetnames:
        for row in wb[name].iter_rows(max_row=min(wb[name].max_row, 80), max_col=15):
            for cell in row:
                if isinstance(cell.value, str) and "#REF!" in cell.value:
                    refs.append((name, cell.coordinate))

    wb.save(OUT)
    OUT_PROJECT.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(OUT, OUT_PROJECT)

    print("FORMATTED", OUT)
    print("SHEETS", len(sheet_order))
    print("IT_SHEETS", len(it_sheets))
    print("BORDER_SHEETS", len(borders_sheets))
    print("REFS", refs or "None")
    print("ORDER_OK", list(wb.sheetnames) == sheet_order)


if __name__ == "__main__":
    main()
