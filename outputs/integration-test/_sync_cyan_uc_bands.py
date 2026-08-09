# -*- coding: utf-8 -*-
"""
Rewrite cyan UC group bands + case rows from uc_modules (SRS titles),
preserving Round 1/2/3 status/date/tester/note per Test Case ID.
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parent))
from uc_modules import MODULES, iter_cases  # noqa: E402

EXCEL = Path(
    r"C:\Users\phong\Downloads\intergration test"
    r"\SEP490_G23_Report5.2_Integration Test_COMPLETED.xlsx"
)
PROJ = Path(r"D:\EngLishLab\EnglishLab\outputs\integration-test")
BACKUP = EXCEL.with_name(EXCEL.stem + "_BEFORE_UC_BAND_SYNC.xlsx")

GROUP_FILL = PatternFill("solid", fgColor="CCFFFF")
GROUP_FONT = Font(name="Tahoma", size=10, bold=True, color="000000")
CASE_FILL = PatternFill("solid", fgColor="FFFFFF")
NORMAL_FONT = Font(name="Tahoma", size=10, color="000000")
THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

STATUS_COLS = (6, 9, 12)
DATE_COLS = (7, 10, 13)
TESTER_COLS = (8, 11, 14)


def read_results(ws) -> dict[str, dict]:
    out: dict[str, dict] = {}
    for row in range(11, min(ws.max_row or 11, 500) + 1):
        cid = ws.cell(row, 1).value
        if not isinstance(cid, str) or not cid.startswith("IT_"):
            continue
        out[cid] = {
            "status": [ws.cell(row, c).value for c in STATUS_COLS],
            "date": [ws.cell(row, c).value for c in DATE_COLS],
            "tester": [ws.cell(row, c).value for c in TESTER_COLS],
            "note": ws.cell(row, 15).value,
        }
    return out


def clear_body(ws, start=11, end=80):
    for r in range(start, end + 1):
        for c in range(1, 16):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.font = NORMAL_FONT
            cell.border = Border()
            cell.alignment = Alignment()
        ws.row_dimensions[r].height = None


def write_module(ws, module, results: dict[str, dict]):
    ws["B2"] = module["function"]
    if module.get("requirement"):
        ws["B3"] = module["requirement"]
    clear_body(ws, 11, max(ws.max_row or 40, 40))
    row = 11
    for kind, gname, _scope, case in iter_cases(module):
        if kind == "GROUP":
            ws.cell(row, 1).value = gname
            for c in range(1, 16):
                cell = ws.cell(row, c)
                cell.fill = GROUP_FILL
                cell.font = GROUP_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")
                if c > 1:
                    cell.value = None
            ws.row_dimensions[row].height = 18
            row += 1
            continue

        ws.cell(row, 1).value = case["id"]
        ws.cell(row, 2).value = case["desc"]
        ws.cell(row, 3).value = case["proc"]
        ws.cell(row, 4).value = case["exp"]
        ws.cell(row, 5).value = case["pre"]
        for c in range(1, 6):
            cell = ws.cell(row, c)
            cell.fill = CASE_FILL
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP
        for c in range(6, 16):
            cell = ws.cell(row, c)
            cell.fill = CASE_FILL
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER if c < 15 else WRAP

        prev = results.get(case["id"], {})
        statuses = prev.get("status") or ["Pending", "Pending", "Pending"]
        dates = prev.get("date") or [None, None, None]
        testers = prev.get("tester") or [None, None, None]
        for i, c in enumerate(STATUS_COLS):
            ws.cell(row, c).value = statuses[i] if i < len(statuses) else "Pending"
        for i, c in enumerate(DATE_COLS):
            ws.cell(row, c).value = dates[i] if i < len(dates) else None
        for i, c in enumerate(TESTER_COLS):
            ws.cell(row, c).value = testers[i] if i < len(testers) else None
        note_cell = ws.cell(row, 15)
        note_cell.value = prev.get("note")
        note_cell.fill = CASE_FILL
        note_cell.font = NORMAL_FONT
        note_cell.border = THIN_BORDER
        note_cell.alignment = WRAP
        ws.row_dimensions[row].height = 60
        row += 1


def main():
    if not EXCEL.exists():
        raise SystemExit(f"Missing {EXCEL}")
    shutil.copy2(EXCEL, BACKUP)
    wb = load_workbook(EXCEL)
    changed = []
    for m in MODULES:
        name = m["sheet"]
        if name not in wb.sheetnames:
            print("SKIP missing", name)
            continue
        ws = wb[name]
        results = read_results(ws)
        before = []
        for r in range(11, min(ws.max_row or 11, 80) + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and not v.startswith("IT_"):
                before.append(v)
        write_module(ws, m, results)
        after = []
        for r in range(11, min(ws.max_row or 11, 80) + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and not v.startswith("IT_"):
                after.append(v)
        if before != after or ws["B2"].value != m["function"]:
            changed.append((name, before, after, m["function"]))
            print(name)
            print("  Feature ->", m["function"])
            print("  bands  :", before, "=>", after)

    try:
        wb.save(EXCEL)
        out = EXCEL
    except PermissionError:
        out = EXCEL.with_name(EXCEL.stem + "_UC_BANDS.xlsx")
        wb.save(out)
        print("LOCKED ->", out)
    shutil.copy2(out, PROJ / out.name)
    print("changed sheets", len(changed), "BACKUP", BACKUP, "WROTE", out)


if __name__ == "__main__":
    main()
