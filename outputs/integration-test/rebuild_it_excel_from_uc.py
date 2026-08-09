# -*- coding: utf-8 -*-
"""
Rebuild SEP490_G23_Report5.2_Integration Test_COMPLETED content from SRS UCs.
Keep workbook form/styles; rewrite Test Cases list + each IT sheet body.
"""
from __future__ import annotations

import shutil
from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink

from uc_modules import MODULES, iter_cases

# Prefer newest user-saved COMPLETED.xlsx under Downloads, then project copy.
_DL_DIR = Path(r"C:\Users\phong\Downloads\intergration test")
_SRC_CANDIDATES = []
if _DL_DIR.exists():
    _SRC_CANDIDATES.extend(
        sorted(
            (
                p
                for p in _DL_DIR.glob("SEP490_G23_Report5.2_Integration Test_COMPLETED*.xlsx")
                if not p.name.startswith("~$")
                and "TIENG_VIET" not in p.name
                and "HONEST" not in p.name
                and "UC_REBUILD" not in p.name
                and "SRS_UC" not in p.name
            ),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
    )
_SRC_CANDIDATES.extend(
    [
        Path(
            r"D:\EngLishLab\EnglishLab\outputs\integration-test"
            r"\SEP490_G23_Report5.2_Integration Test_COMPLETED.xlsx"
        ),
        Path(
            r"D:\EngLishLab\EnglishLab\outputs\integration-test"
            r"\SEP490_G23_Report5.2_Integration Test_COMPLETED_HONEST_FORMATTED.xlsx"
        ),
    ]
)
SRC = next((p for p in _SRC_CANDIDATES if p.exists()), _SRC_CANDIDATES[-1])
OUT_PROJ = Path(
    r"D:\EngLishLab\EnglishLab\outputs\integration-test"
    r"\SEP490_G23_Report5.2_Integration Test_COMPLETED.xlsx"
)
OUT_DL = Path(
    r"C:\Users\phong\Downloads\intergration test"
    r"\SEP490_G23_Report5.2_Integration Test_COMPLETED_UC_REBUILD.xlsx"
)
OUT_DL_ALT = Path(
    r"C:\Users\phong\Downloads\intergration test"
    r"\SEP490_G23_Report5.2_Integration Test_COMPLETED.xlsx"
)

ISSUE_DATE = "2026-08-06"
VERSION = "v6.0-IT-sample"
CREATOR = "phongdx"
TESTER = "phongdx"

LINK_FONT = Font(name="Tahoma", size=10, color="0000FF", underline="single")
# School-sample IT sheet colors (green header + cyan body)
HEADER_FILL = PatternFill("solid", fgColor="76923C")
HEADER_FONT = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")
GROUP_FILL = PatternFill("solid", fgColor="CCFFFF")  # pale cyan UC group band only
GROUP_FONT = Font(name="Tahoma", size=10, bold=True, color="000000")
CASE_FILL = PatternFill("solid", fgColor="FFFFFF")  # case rows stay white
NORMAL_FONT = Font(name="Tahoma", size=10, color="000000")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
NO_FILL = PatternFill(fill_type=None)
THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

STATUS_COLS = (6, 9, 12)
DATE_COLS = (7, 10, 13)
TESTER_COLS = (8, 11, 14)
FILL_ROUNDS = True  # sample Round 1 = Pending; leave date/tester blank until run

IT_HEADER_LABELS = [
    "Test Case ID",
    "Test Case Description",
    "Test Case Procedure",
    "Expected Results",
    "Pre-conditions",
    "Round 1",
    "Test date",
    "Tester",
    "Round 2",
    "Test date",
    "Tester",
    "Round 3",
    "Test date",
    "Tester",
    "Note",
]


def set_sheet_hyperlink(cell, sheet_name: str, display: str | None = None):
    display = display or sheet_name
    cell.value = display
    cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{sheet_name}'!A1", display=display)
    cell.font = LINK_FONT


def clear_rows(ws, start_row: int, end_row: int, max_col: int = 14):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.hyperlink = None
            cell.fill = NO_FILL
            cell.border = Border()


def style_it_header_row(ws, row: int = 10):
    for c, label in enumerate(IT_HEADER_LABELS, start=1):
        cell = ws.cell(row, c)
        cell.value = label
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def write_module_sheet(ws, module: dict):
    # Feature = UC function name (no Controller)
    ws["A2"] = "Feature"
    ws["B2"] = module["function"]
    ws["A3"] = "Test requirement"
    ws["B3"] = module["requirement"]
    # keep formulas in A4/B4 and rounds if present
    if ws["A4"].value is None:
        ws["A4"] = "Number of TCs"
    ws["B4"] = '=COUNTIF(A11:A1000,"IT_*")'

    # Ensure round labels remain English for COUNTIF compatibility with Passed/Failed
    ws["A5"] = "Testing Round"
    ws["B5"] = "Passed"
    ws["C5"] = "Failed"
    ws["D5"] = "Pending"
    ws["E5"] = "N/A"

    style_it_header_row(ws, 10)

    end_clear = min(ws.max_row or 200, 200)
    clear_rows(ws, 11, end_clear, 15)
    for r in range(11, end_clear + 1):
        ws.row_dimensions[r].height = None

    row = 11
    for kind, gname, scope, case in iter_cases(module):
        if kind == "GROUP":
            # School sample: cyan UC group band (include Note column O)
            ws.cell(row, 1).value = gname
            for c in range(1, 16):
                cell = ws.cell(row, c)
                cell.fill = GROUP_FILL
                cell.font = GROUP_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")
                if c > 1:
                    cell.value = None
            ws.row_dimensions[row].height = 18
            row += 1
            continue

        ws.cell(row, 1).value = case["id"]
        ws.cell(row, 2).value = case["desc"]
        ws.cell(row, 3).value = case["proc"]
        ws.cell(row, 4).value = case["exp"]
        ws.cell(row, 5).value = case["pre"]
        if FILL_ROUNDS:
            status = case.get("status", "Pending")
            for c in STATUS_COLS:
                ws.cell(row, c).value = status
            for c in DATE_COLS + TESTER_COLS:
                ws.cell(row, c).value = None
        else:
            for c in STATUS_COLS + DATE_COLS + TESTER_COLS:
                ws.cell(row, c).value = None

        for c in range(1, 16):
            cell = ws.cell(row, c)
            cell.fill = CASE_FILL
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if c <= 5 or c == 15:
                cell.alignment = WRAP
            else:
                cell.alignment = Alignment(horizontal="center", vertical="top")
        # wrap-friendly height; avoid leftover huge cloned row heights
        ws.row_dimensions[row].height = 75
        row += 1
    # do not append SRS mapping footer


def rewrite_test_cases_sheet(ws):
    # Keep title / project refs / environment; rewrite function table from row 8
    ws["D1"] = "TEST CASE LIST"
    # environment already there; update lightly
    env = (
        "1. Server: Spring Boot 4 + JDK 21 (@SpringBootTest + MockMvc)\n"
        "2. Database: PostgreSQL with migrations applied\n"
        "3. Web Browser: N/A for API Integration Tests (optional Postman/curl)\n"
        "4. JWT security enabled; roles seeded\n"
        "5. External PayOS / Mail / Gemini / Meet stubbed or sandbox\n"
        "6. Feature names mapped to SRS Report3 Use Cases (UC-xx titles, no Controller suffix)"
    )
    ws["D5"] = env

    # header row 8
    ws["B8"] = "No"
    ws["C8"] = "Function Name"
    ws["D8"] = "Sheet Name"
    ws["E8"] = "Description"
    ws["F8"] = "Pre-Condition"

    # clear old module rows
    clear_rows(ws, 9, 80, 8)

    for i, m in enumerate(MODULES, 1):
        r = 8 + i
        ws.cell(r, 2).value = i
        ws.cell(r, 3).value = m["function"]  # UC title style
        set_sheet_hyperlink(ws.cell(r, 4), m["sheet"], m["sheet"])
        ucs = ", ".join(m.get("ucs", []))
        n_cases = sum(1 for k, *_ , c in iter_cases(m) if k == "CASE")
        ws.cell(r, 5).value = (
            f"{m['requirement']}\n"
            f"Use cases: {ucs}.\n"
            f"Components: {m.get('components','')}.\n"
            f"Test cases in sheet: {n_cases}."
        )
        ws.cell(r, 5).alignment = WRAP
        ws.cell(r, 6).value = (
            f"DB up; roles seeded; target sheet `{m['sheet']}` exists; "
            "external systems stubbed as needed for this module."
        )
        ws.cell(r, 6).alignment = WRAP


def rewrite_statistics(ws):
    """Rebuild Test Statistics to school-sample layout with live formulas to IT_* sheets."""
    header_fill = PatternFill("solid", fgColor="000080")
    header_font = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")
    value_font = Font(name="Tahoma", size=10)
    label_font = Font(name="Tahoma", size=10, bold=True)
    cov_font = Font(name="Tahoma", size=10, color="993300")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    # Title + metadata (sample layout)
    ws["B1"] = "TEST STATISTICS"
    ws["B3"] = "Project Name"
    ws["C3"] = "=Cover!B4"
    ws["E3"] = "Creator"
    ws["F3"] = CREATOR
    ws["B4"] = "Project Code"
    ws["C4"] = "=Cover!B5"
    ws["E4"] = "Reviewer/Approver"
    ws["B5"] = "Document Code"
    ws["C5"] = '=C4&"_"&"Test Report"&"_"&"vx.x"'
    ws["E5"] = "Issue Date"
    ws["H5"] = ISSUE_DATE
    ws["B6"] = "Notes"
    codes = ", ".join(m["code"] for m in MODULES)
    ws["C6"] = (
        f"Release Integration Test pack includes {len(MODULES)} modules: {codes}. "
        "Module code = Feature name from each IT_* sheet (SRS Use Case title). "
        "Passed/Failed/Pending/N/A pulled from Round 1 counts on each sheet."
    )
    ws["C6"].alignment = WRAP
    ws["C6"].font = value_font
    for addr in ("C3", "C4", "C5", "F3", "H5"):
        ws[addr].font = value_font
    for addr in ("B3", "B4", "B5", "B6", "E3", "E4", "E5"):
        ws[addr].font = label_font

    # Clear old module / subtotal / coverage block (keep header row 10)
    for r in range(11, 80):
        for c in range(2, 9):
            cell = ws.cell(r, c)
            cell.value = None
            cell.hyperlink = None
            cell.fill = PatternFill()
            cell.font = value_font
            cell.border = Border()

    # Table header (sample)
    headers = [
        (2, "No"),
        (3, "Module code"),
        (4, "Passed"),
        (5, "Failed"),
        (6, "Pending"),
        (7, "N/A"),
        (8, "Number of  test cases"),
    ]
    for c, text in headers:
        cell = ws.cell(10, c)
        cell.value = text
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin

    # One row per current MODULE, formulas to IT_* sheets (no broken external refs)
    for i, m in enumerate(MODULES, start=1):
        r = 10 + i
        sheet = m["sheet"]
        ws.cell(r, 2).value = i
        ws.cell(r, 3).value = f"='{sheet}'!B2"  # Feature / Module code
        ws.cell(r, 4).value = f"='{sheet}'!B6"  # Passed Round 1
        ws.cell(r, 5).value = f"='{sheet}'!C6"  # Failed
        ws.cell(r, 6).value = f"='{sheet}'!D6"  # Pending
        ws.cell(r, 7).value = f"='{sheet}'!E6"  # N/A
        ws.cell(r, 8).value = f"='{sheet}'!B4"  # Number of TCs
        for c in range(2, 9):
            cell = ws.cell(r, c)
            cell.font = value_font
            cell.border = thin
            cell.hyperlink = None
            if c == 2:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c >= 4:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "0"

    last = 10 + len(MODULES)
    sub_r = last + 2
    # Sub total band
    for c in range(2, 9):
        cell = ws.cell(sub_r, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
        cell.value = None
    ws.cell(sub_r, 3).value = "Sub total"
    ws.cell(sub_r, 4).value = f"=SUM(D11:D{last})"
    ws.cell(sub_r, 5).value = f"=SUM(E11:E{last})"
    ws.cell(sub_r, 6).value = f"=SUM(F11:F{last})"
    ws.cell(sub_r, 7).value = f"=SUM(G11:G{last})"
    ws.cell(sub_r, 8).value = f"=SUM(H11:H{last})"
    for c in range(4, 9):
        ws.cell(sub_r, c).number_format = "0"
        ws.cell(sub_r, c).alignment = Alignment(horizontal="center", vertical="center")

    cov_r = sub_r + 2
    ws.cell(cov_r, 3).value = "Test coverage"
    ws.cell(cov_r, 3).font = cov_font
    ws.cell(cov_r, 5).value = (
        f"=IF((H{sub_r}-G{sub_r})=0,0,(D{sub_r}+E{sub_r})*100/(H{sub_r}-G{sub_r}))"
    )
    ws.cell(cov_r, 5).number_format = "0.00"
    ws.cell(cov_r, 5).font = cov_font
    ws.cell(cov_r, 6).value = "%"
    ws.cell(cov_r, 6).font = cov_font

    ws.cell(cov_r + 1, 3).value = "Test successful coverage"
    ws.cell(cov_r + 1, 3).font = cov_font
    ws.cell(cov_r + 1, 5).value = f"=IF((H{sub_r}-G{sub_r})=0,0,D{sub_r}*100/(H{sub_r}-G{sub_r}))"
    ws.cell(cov_r + 1, 5).number_format = "0.00"
    ws.cell(cov_r + 1, 5).font = cov_font
    ws.cell(cov_r + 1, 6).value = "%"
    ws.cell(cov_r + 1, 6).font = cov_font


def rewrite_cover(ws):
    ws["B2"] = "TEST REPORT DOCUMENT"
    ws["B4"] = "EnglishLab"
    ws["B5"] = "SEP490_G23"
    ws["F4"] = CREATOR
    ws["F5"] = ISSUE_DATE
    ws["F6"] = VERSION
    # change log
    ws["A11"] = ISSUE_DATE
    ws["B11"] = VERSION
    ws["C11"] = "Rewrite all IT cases to Controller-Service-Repository sample style (MockMvc, no UI)"
    ws["D11"] = "M"
    ws["E11"] = (
        "Function Name = SRS Use Case title. "
        "Sheet Name = Test Case ID prefix (e.g. IT_COURSE for IT_COURSE_01). "
        "Procedure = Call API via MockMvc; Controller delegates to Service; Service uses Repository. "
        "Expected = HTTP status and JSON/DB checks. Round 1 = Pending."
    )
    ws["F11"] = "SRS Report3 Use Cases (UC-01..UC-51); existing IT form"


def reorder_it_sheets(wb):
    """Move IT sheets to match MODULES priority after Cover/Test Cases/Statistics."""
    fixed = ["Cover", "Test Cases", "Test Statistics"]
    desired = fixed + [m["sheet"] for m in MODULES]
    # keep any leftover sheets at end
    leftovers = [n for n in wb.sheetnames if n not in desired]
    order = [n for n in desired if n in wb.sheetnames] + leftovers
    for idx, name in enumerate(order):
        current = wb.sheetnames.index(name)
        if current != idx:
            wb.move_sheet(name, offset=idx - current)


def clone_it_sheet(wb, target: str):
    """Clone layout from any existing IT_* sheet when a new UC sheet is needed."""
    src_name = next((n for n in wb.sheetnames if n.startswith("IT_")), None)
    if src_name is None:
        raise SystemExit(f"Cannot clone sheet {target!r}: no IT_* template in workbook")
    ws = wb.copy_worksheet(wb[src_name])
    ws.title = target
    print(f"CLONED {src_name} -> {target}")
    return ws


def main():
    if not SRC.exists():
        raise SystemExit(f"Missing template: {SRC}")

    OUT_PROJ.parent.mkdir(parents=True, exist_ok=True)
    OUT_DL.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SRC, OUT_PROJ)
    wb = load_workbook(OUT_PROJ)

    # Rename template / previous sheets to IT_* (same as Test Case ID prefix)
    try:
        from _sheet_legacy_map import LEGACY_ALIASES
    except ImportError:
        LEGACY_ALIASES = {}
    # Build code -> target from MODULES
    for m in MODULES:
        target = m["sheet"]
        if len(target) > 31:
            raise SystemExit(f"Sheet name > 31 chars: {target!r}")
        # find any existing alias sheet for this module
        candidates = [old for old, new in LEGACY_ALIASES.items() if new == target]
        candidates.append(target)
        found = next((n for n in candidates if n in wb.sheetnames), None)
        if found is None:
            clone_it_sheet(wb, target)
            continue
        if found != target:
            if target in wb.sheetnames and target != found:
                wb.remove(wb[target])
            wb[found].title = target

    rewrite_cover(wb["Cover"])
    rewrite_test_cases_sheet(wb["Test Cases"])
    rewrite_statistics(wb["Test Statistics"])

    for m in MODULES:
        name = m["sheet"]
        if name not in wb.sheetnames:
            clone_it_sheet(wb, name)
        write_module_sheet(wb[name], m)
        # clear any leftover "SRS mapping" rows below cases
        for r in range(11, min(wb[name].max_row or 200, 200) + 1):
            if wb[name].cell(r, 1).value == "SRS mapping":
                wb[name].cell(r, 1).value = None
                wb[name].cell(r, 2).value = None
        print(f"OK {name} :: {m['function']} :: cases="
              f"{sum(1 for k, *_ , c in iter_cases(m) if k == 'CASE')}")


    reorder_it_sheets(wb)

    # Drop sheets that are no longer in MODULES (no SRS UC description)
    keep = {"Cover", "Test Cases", "Test Statistics"} | {m["sheet"] for m in MODULES}
    for name in list(wb.sheetnames):
        if name not in keep:
            wb.remove(wb[name])
            print(f"REMOVED sheet {name}")

    wb.save(OUT_PROJ)
    OUT_DL.parent.mkdir(parents=True, exist_ok=True)
    for dest in (OUT_DL, OUT_DL_ALT):
        try:
            shutil.copy2(OUT_PROJ, dest)
            print(f"OUT {dest}")
            continue
        except PermissionError:
            pass
        for suffix in ("_NEW", "_SRS_UC", "_LATEST"):
            alt = dest.with_name(dest.stem + f"{suffix}.xlsx")
            try:
                shutil.copy2(OUT_PROJ, alt)
                print(f"LOCKED {dest.name} -> {alt.name}")
                break
            except PermissionError:
                continue
        else:
            print(f"LOCKED all copies for {dest.name}; kept {OUT_PROJ}")
    print(f"OUT {OUT_PROJ}")


if __name__ == "__main__":
    main()
