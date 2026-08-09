# -*- coding: utf-8 -*-
from openpyxl import load_workbook

p = r"C:\Users\phong\Downloads\intergration test\SEP490_G23_Report5.2_Integration Test_COMPLETED_v4_FORMATTED.xlsx"
wb = load_workbook(p)
st = wb["Test Statistics"]
links = [h for h in (st._hyperlinks or []) if str(getattr(h, "ref", "")).upper().startswith("C")]
print("module-code hyperlinks", links or "None")
for r in (11, 12, 34, 36, 38, 39):
    print(r, [st.cell(r, c).value for c in range(2, 9)])
print("Auth B2/B4/B6", wb["IT - Auth"]["B2"].value, wb["IT - Auth"]["B4"].value, wb["IT - Auth"]["B6"].value)
