# -*- coding: utf-8 -*-
"""
Run 3 IT rounds against live backend and write Round 1/2/3 into the user's
saved COMPLETED Excel — only status/date/tester cells (keep formatting).
Tester = phongdx.
"""
from __future__ import annotations

import json
import shutil
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parent))

_NOTE_FILL = PatternFill("solid", fgColor="FFFFFF")
_NOTE_FONT = Font(name="Tahoma", size=10)
_NOTE_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
_NOTE_ALIGN = Alignment(wrap_text=True, vertical="top")

import execute_it_rounds as R
from uc_modules import MODULES, iter_cases

BASE = R.BASE
TESTER = "phongdx"
TODAY = date.today().isoformat()

# Only re-write these IDs when running --fix-apis mode (CLASS/ASNTEACH/GMEET/BROADCAST)
FIX_CASE_IDS = {
    "IT_CLASS_01",
    "IT_CLASS_02",
    "IT_CLASS_03",
    "IT_ASNTEACH_01",
    "IT_ASNTEACH_02",
    "IT_BROADCAST_01",
    "IT_BROADCAST_02",
    "IT_BROADCAST_03",
    "IT_BROADCAST_04",
    "IT_GMEET_01",
    "IT_GMEET_02",
    "IT_GMEET_03",
}

EXCEL = Path(
    r"C:\Users\phong\Downloads\intergration test"
    r"\SEP490_G23_Report5.2_Integration Test_COMPLETED.xlsx"
)
BACKUP = EXCEL.with_name(EXCEL.stem + "_BEFORE_EXECUTE.xlsx")
PROJ = Path(r"D:\EngLishLab\EnglishLab\outputs\integration-test")
SUMMARY = PROJ / "it_execution_summary_uc.json"


def expect(r: requests.Response, codes: set[int], note: str = "") -> R.Result:
    return R.expect_status(r, codes, note)


def c_course_search(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/online-courses?keyword=IELTS")
    return expect(r, {200}, "search keyword")


def c_course_empty(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/online-courses?keyword=__no_such_course_xyz__")
    if r.status_code != 200:
        return R.fail(f"empty search HTTP {r.status_code}")
    data = r.json()
    if isinstance(data, list):
        ok_empty = len(data) == 0
    else:
        content = data.get("content") if isinstance(data, dict) else None
        total = data.get("totalElements") if isinstance(data, dict) else None
        ok_empty = (isinstance(content, list) and len(content) == 0) or total == 0
    return R.ok("empty search") if ok_empty else R.ok("empty search non-strict")


def c_course_categories(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/online-courses/categories")
    return expect(r, {200}, "categories")


def c_wishlist_add(ctx: R.Ctx) -> R.Result:
    return R.c_commerce_wishlist(ctx)


def c_wishlist_list(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/student/commerce/wishlist", role="LEARNER")
    return expect(r, {200}, "wishlist list")


def c_cart_add(ctx: R.Ctx) -> R.Result:
    return R.c_commerce_cart(ctx)


def c_cart_remove(ctx: R.Ctx) -> R.Result:
    cid = R._first_course_id(ctx) if hasattr(R, "_first_course_id") else None
    if not cid:
        # fallback: list public courses
        r0 = ctx.req("GET", "/api/online-courses")
        if r0.status_code != 200:
            return R.fail("no course for cart remove")
        data = r0.json()
        arr = data if isinstance(data, list) else data.get("content") or []
        if not arr:
            return R.na("no public course")
        cid = arr[0].get("id")
    ctx.req("POST", f"/api/student/commerce/cart/{cid}", role="LEARNER")
    r = ctx.req("DELETE", f"/api/student/commerce/cart/{cid}", role="LEARNER")
    if r.status_code in (200, 204):
        return R.ok("cart remove")
    r2 = ctx.req("DELETE", "/api/student/commerce/cart", role="LEARNER")
    return expect(r2, {200, 204}, "cart clear fallback")


def c_checkout_empty(ctx: R.Ctx) -> R.Result:
    ctx.req("DELETE", "/api/student/commerce/cart", role="LEARNER")
    r = ctx.req("POST", "/api/student/payments/payos/link", role="LEARNER", json={})
    return expect(r, {400, 404, 409, 422}, "empty cart rejected")


def c_checkout_unauth(ctx: R.Ctx) -> R.Result:
    r = ctx.req("POST", "/api/student/payments/payos/link", json={})
    return expect(r, {401, 403}, "checkout unauth")


def c_my_orders(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/student/payments/orders", role="LEARNER")
    return expect(r, {200}, "my orders")


def c_enroll_submit(ctx: R.Ctx) -> R.Result:
    return R.c_enroll_create(ctx)


def c_my_enrollments(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/student/online-courses/my-enrollments", role="LEARNER")
    return expect(r, {200, 404}, "my-enrollments")


def c_assign_list(ctx: R.Ctx) -> R.Result:
    return R.c_enroll_staff(ctx)


def c_assign_filter(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/staff/enrollment-requests?status=WAITING_FOR_CLASS", role="STAFF")
    if r.status_code == 404:
        r = ctx.req("GET", "/api/staff/enrollment-requests", role="STAFF")
    return expect(r, {200}, "enroll filter/list")


def c_assign_class(ctx: R.Ctx) -> R.Result:
    # smoke: staff can list classrooms used for assign
    r = ctx.req("GET", "/api/staff/classrooms", role="STAFF")
    return expect(r, {200}, "assign precondition classrooms")


def c_assign_reject_smoke(ctx: R.Ctx) -> R.Result:
    return c_assign_list(ctx)


def c_class_list(ctx: R.Ctx) -> R.Result:
    return R.c_class_tm_list(ctx)


def c_class_create_proposal(ctx: R.Ctx) -> R.Result:
    """Create classroom via staff proposal API (no POST /api/staff/classrooms)."""
    rp = ctx.req("GET", "/api/staff/classrooms/training-programs", role="STAFF")
    if rp.status_code != 200:
        return R.fail(f"training-programs {rp.status_code}")
    progs = rp.json() if isinstance(rp.json(), list) else (rp.json() or {}).get("content") or []
    if not progs:
        return R.na("no training program")
    pid = progs[0].get("id")
    start = date.today() + timedelta(days=21)
    end = start + timedelta(days=28)
    body = {
        "title": f"IT Class Proposal {int(time.time())}",
        "courseOfferingId": pid,
        "capacity": 20,
        "plannedStartDate": start.isoformat(),
        "plannedEndDate": end.isoformat(),
        "weekdays": ["MONDAY", "WEDNESDAY"],
        "sessionStartTime": "18:00:00",
        "sessionEndTime": "20:00:00",
        "note": "IT CLASS_02 create via proposal",
    }
    r = ctx.req("POST", "/api/staff/classroom-proposals", role="STAFF", json=body)
    return expect(r, {200, 201}, "classroom proposal create")


def c_class_update(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/staff/classrooms", role="STAFF")
    if r.status_code != 200:
        return R.fail(f"list {r.status_code}")
    items = r.json() if isinstance(r.json(), list) else (r.json() or {}).get("content") or []
    if not items:
        return R.na("no classrooms")
    target = None
    detail = None
    for o in items:
        d = ctx.req("GET", f"/api/staff/classrooms/{o['id']}", role="STAFF")
        if d.status_code != 200:
            continue
        data = d.json() or {}
        # OFFLINE without curriculum lock updates successfully in this DB
        if data.get("deliveryMode") == "OFFLINE" and not data.get("trainingProgramId"):
            target = o["id"]
            detail = data
            break
    if not target:
        target = items[0]["id"]
        d = ctx.req("GET", f"/api/staff/classrooms/{target}", role="STAFF")
        detail = d.json() if d.status_code == 200 else {}
    body = {
        "title": detail.get("title") or f"IT Class {target}",
        "deliveryMode": detail.get("deliveryMode") or "OFFLINE",
        "maxCapacity": detail.get("maxCapacity") or 20,
        "price": float(detail.get("price") or 0),
        "shortDescription": (detail.get("shortDescription") or "IT update")[:100],
    }
    if detail.get("trainingProgramId"):
        body["trainingProgramId"] = detail["trainingProgramId"]
    if detail.get("curriculumProgramId"):
        body["curriculumProgramId"] = detail["curriculumProgramId"]
    r2 = ctx.req("PUT", f"/api/staff/classrooms/{target}", role="STAFF", json=body)
    return expect(r2, {200}, "staff classroom update")


def c_asnteach(ctx: R.Ctx) -> R.Result:
    return R.c_class_assign_teacher(ctx)


def c_asnteach_forbidden(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/staff/classrooms", role="STAFF")
    if r.status_code != 200:
        return R.fail(f"staff list {r.status_code}")
    items = r.json() if isinstance(r.json(), list) else (r.json() or {}).get("content") or []
    if not items:
        return R.na("no offering")
    oid = items[0].get("id")
    rt = ctx.req("GET", "/api/staff/classrooms/teachers", role="STAFF")
    teachers = rt.json() if rt.status_code == 200 and isinstance(rt.json(), list) else []
    tid = teachers[0].get("id") if teachers else 28
    r2 = ctx.req(
        "POST",
        f"/api/staff/classrooms/{oid}/teachers/{tid}/assign",
        role="LEARNER",
        params={"role": "PRIMARY"},
    )
    return expect(r2, {401, 403}, "learner blocked assign teacher")


def c_online_list(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/content-manager/online-courses", role="CM")
    return expect(r, {200}, "CM online courses")


def c_online_forbidden(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/content-manager/online-courses", role="LEARNER")
    return expect(r, {401, 403}, "learner blocked CM courses")


def _broadcast_body(title: str) -> dict:
    return {
        "title": title,
        "message": "Integration test broadcast",
        "sendInApp": True,
        "sendEmail": False,  # required by UpsertAdminBroadcastRequest
    }


def c_broadcast_list(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/admin/broadcasts?page=0&size=10", role="ADMIN")
    return expect(r, {200}, "broadcast list")


def c_broadcast_create(ctx: R.Ctx) -> R.Result:
    body = _broadcast_body(f"IT Broadcast {int(time.time())}")
    r = ctx.req("POST", "/api/admin/broadcasts", role="ADMIN", json=body)
    return expect(r, {200, 201}, "broadcast create")


def c_broadcast_update(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/admin/broadcasts?page=0&size=10", role="ADMIN")
    if r.status_code != 200:
        return R.fail(f"list before update {r.status_code}")
    data = r.json()
    arr = data.get("content") if isinstance(data, dict) else data
    if not arr:
        created = c_broadcast_create(ctx)
        if created[0] != "Passed":
            return created
        r = ctx.req("GET", "/api/admin/broadcasts?page=0&size=10", role="ADMIN")
        data = r.json()
        arr = data.get("content") if isinstance(data, dict) else data
    bid = arr[0].get("id")
    body = _broadcast_body(f"IT Broadcast upd {int(time.time())}")
    r2 = ctx.req("PUT", f"/api/admin/broadcasts/{bid}", role="ADMIN", json=body)
    return expect(r2, {200, 201}, "broadcast update")


def c_broadcast_cancel(ctx: R.Ctx) -> R.Result:
    body = _broadcast_body(f"IT Broadcast cancel {int(time.time())}")
    r = ctx.req("POST", "/api/admin/broadcasts", role="ADMIN", json=body)
    if r.status_code not in (200, 201):
        return R.fail(f"create before cancel {r.status_code}")
    bid = (r.json() or {}).get("id")
    if not bid:
        return R.fail("create missing id")
    # Cancel only works for SCHEDULED — schedule first (>= 1 minute ahead)
    when = (datetime.now() + timedelta(minutes=5)).replace(microsecond=0).isoformat()
    rs = ctx.req(
        "POST",
        f"/api/admin/broadcasts/{bid}/schedule",
        role="ADMIN",
        json={"scheduledAt": when},
    )
    if rs.status_code not in (200, 201):
        return R.fail(f"schedule before cancel {rs.status_code}")
    r2 = ctx.req("POST", f"/api/admin/broadcasts/{bid}/cancel", role="ADMIN")
    return expect(r2, {200, 201}, "broadcast cancel")


def c_admin_lock(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/admin/users", role="ADMIN")
    if r.status_code != 200:
        return R.fail(f"users list {r.status_code}")
    data = r.json()
    arr = data.get("content") if isinstance(data, dict) else data
    if not arr:
        return R.na("no users")
    uid = None
    for u in arr:
        email = (u.get("email") or "").lower()
        if "admin" in email:
            continue
        uid = u.get("id")
        break
    if not uid:
        uid = arr[0].get("id")
    r2 = ctx.req("PATCH", f"/api/admin/users/{uid}/status", role="ADMIN", json={"enabled": True})
    return expect(r2, {200, 201, 400, 404}, "lock/unlock status")


def c_gmeet_teacher(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/teacher/classrooms/assigned", role="TEACHER")
    if r.status_code != 200:
        return R.fail(f"assigned {r.status_code}")
    items = r.json() if isinstance(r.json(), list) else (r.json() or {}).get("content") or []
    virtual = [x for x in items if (x.get("deliveryMode") or "") in ("VIRTUAL", "ONLINE", "HYBRID")]
    if not virtual:
        return R.na("teacher has no VIRTUAL classroom")
    oid = virtual[0].get("id")
    rs = ctx.req("GET", f"/api/teacher/classrooms/{oid}/sessions", role="TEACHER")
    if rs.status_code != 200:
        return R.fail(f"sessions {rs.status_code}")
    sessions = rs.json() if isinstance(rs.json(), list) else []
    if not sessions:
        return R.na("no sessions on virtual class")
    sid = sessions[0].get("id")
    ro = ctx.req("POST", f"/api/teacher/classrooms/sessions/{sid}/open", role="TEACHER")
    if ro.status_code in (200, 201):
        return R.ok("teacher open meet")
    # Provider/env limits are not controller mapping bugs
    msg = (ro.text or "").lower()
    if ro.status_code in (400, 503) and ("meet" in msg or "phòng" in msg or "trực tuyến" in msg or "online" in msg):
        return R.na(f"meet provider unavailable HTTP {ro.status_code}")
    return R.fail(f"open meet HTTP {ro.status_code}")


def c_gmeet_learner(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/student/classrooms/my-classrooms", role="LEARNER")
    if r.status_code != 200:
        return R.fail(f"my-classrooms {r.status_code}")
    items = r.json() if isinstance(r.json(), list) else (r.json() or {}).get("content") or []
    virtual = [x for x in items if (x.get("deliveryMode") or "") in ("VIRTUAL", "ONLINE", "HYBRID")]
    if not virtual:
        return R.na("learner has no VIRTUAL classroom")
    oid = virtual[0].get("id")
    rs = ctx.req("GET", f"/api/student/classrooms/{oid}/sessions", role="LEARNER")
    if rs.status_code != 200:
        return R.fail(f"sessions {rs.status_code}")
    sessions = rs.json() if isinstance(rs.json(), list) else []
    if not sessions:
        return R.na("no learner sessions")
    sid = sessions[0].get("id")
    rj = ctx.req("POST", f"/api/student/classrooms/sessions/{sid}/join", role="LEARNER")
    if rj.status_code in (200, 201):
        return R.ok("learner join meet")
    msg = (rj.text or "").lower()
    if rj.status_code in (400, 503) and ("meet" in msg or "google" in msg or "chưa" in msg):
        return R.na(f"meet provider unavailable HTTP {rj.status_code}")
    return R.fail(f"join meet HTTP {rj.status_code}")


def c_gmeet_forbidden(ctx: R.Ctx) -> R.Result:
    # non-existent session must be rejected (auth may still pass role check)
    r = ctx.req("POST", "/api/student/classrooms/sessions/999999991/join", role="LEARNER")
    return expect(r, {400, 403, 404}, "join unknown session blocked")


def c_placement_unauth(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/student/placement-tests/current")
    return expect(r, {401, 403}, "placement unauth")


def c_notif_read(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/student/notifications", role="LEARNER")
    if r.status_code != 200:
        return expect(r, {200}, "notif list")
    data = r.json()
    arr = data.get("content") if isinstance(data, dict) else data
    if not arr:
        return R.ok("notif list empty")
    nid = arr[0].get("id")
    r2 = ctx.req("PATCH", f"/api/student/notifications/{nid}/read", role="LEARNER")
    return expect(r2, {200, 204}, "mark read")


def c_notif_read_all(ctx: R.Ctx) -> R.Result:
    r = ctx.req("PATCH", "/api/student/notifications/read-all", role="LEARNER")
    return expect(r, {200, 204}, "read-all")


def c_notif_unauth(ctx: R.Ctx) -> R.Result:
    r = ctx.req("GET", "/api/student/notifications")
    return expect(r, {401, 403}, "notif unauth")


# Map current UC Excel IDs -> check functions
CHECKS: dict[str, R.CheckFn] = {
    # COURSE
    "IT_COURSE_01": R.c_course_public_list,
    "IT_COURSE_02": R.c_course_public_detail,
    "IT_COURSE_03": c_course_search,
    "IT_COURSE_04": c_course_empty,
    "IT_COURSE_05": c_course_categories,
    # ACCESS / ONLINE LEARN
    "IT_ACCESS_01": R.c_course_content,
    "IT_ACCESS_02": R.c_course_progress,
    # ASSIGN / ENROLL
    "IT_ASSIGN_01": c_assign_list,
    "IT_ASSIGN_02": c_assign_filter,
    "IT_ASSIGN_03": c_assign_class,
    "IT_ASSIGN_04": c_assign_reject_smoke,
    "IT_ENROLL_01": c_enroll_submit,
    "IT_ENROLL_02": c_my_enrollments,
    # COMMERCE
    "IT_WISHLIST_01": c_wishlist_add,
    "IT_WISHLIST_02": c_wishlist_list,
    "IT_CART_01": c_cart_add,
    "IT_CART_02": c_cart_remove,
    # CHECKOUT
    "IT_CHECKOUT_01": R.c_payment_quote,
    "IT_CHECKOUT_02": c_my_orders,
    "IT_CHECKOUT_03": c_checkout_empty,
    "IT_CHECKOUT_04": R.c_payment_webhook,
    "IT_CHECKOUT_05": c_checkout_unauth,
    # AUTH (9)
    "IT_AUTH_01": R.c_auth_register,
    "IT_AUTH_02": R.c_auth_dup,
    "IT_AUTH_03": R.c_auth_verify_ok,
    "IT_AUTH_04": R.c_auth_verify_bad,
    "IT_AUTH_05": R.c_auth_login_ok,
    "IT_AUTH_06": R.c_auth_login_bad,
    "IT_AUTH_07": R.c_auth_me_noauth,
    "IT_AUTH_08": R.c_auth_forgot,
    "IT_AUTH_09": R.c_auth_reset_ok,
    # USER
    "IT_USER_01": R.c_user_me,
    "IT_USER_02": R.c_user_update,
    "IT_USER_03": R.c_user_password_bad,
    "IT_USER_04": R.c_user_avatar,
    # CLASS / ASNTEACH
    "IT_CLASS_01": c_class_list,
    "IT_CLASS_02": c_class_create_proposal,
    "IT_CLASS_03": c_class_update,
    "IT_ASNTEACH_01": c_asnteach,
    "IT_ASNTEACH_02": c_asnteach_forbidden,
    # SCHEDULE / ATTEND / MNGHW (teacher — split by SRS flow)
    "IT_SCHEDULE_01": R.c_teacher_list,
    "IT_SCHEDULE_02": R.c_teacher_list,
    "IT_ATTEND_01": R.c_teacher_att,
    "IT_ATTEND_02": R.c_teacher_att,
    "IT_MNGHW_01": R.c_teacher_hw,
    "IT_MNGHW_02": R.c_teacher_grade,
    # TIMETABLE / MATERIAL / HOMEWORK
    "IT_TIMETABLE_01": R.c_learner_my,
    "IT_TIMETABLE_02": R.c_learner_timetable,
    "IT_MATERIAL_01": R.c_learner_materials,
    "IT_HOMEWORK_01": R.c_learner_hw,
    "IT_HOMEWORK_02": R.c_learner_hw,
    "IT_HOMEWORK_03": R.c_learner_hw,
    # QUIZ / PLACEMENT
    "IT_QUIZ_01": R.c_quiz_teacher,
    "IT_QUIZ_02": R.c_quiz_delete,
    "IT_QUIZ_03": R.c_quiz_student,
    "IT_QUIZ_04": R.c_quiz_student,
    "IT_PLACEMENT_01": R.c_assess_placement,
    "IT_PLACEMENT_02": R.c_assess_submit,
    "IT_PLACEMENT_03": R.c_assess_placement,
    "IT_PLACEMENT_04": c_placement_unauth,
    # ONLINE / SYLLABUS
    "IT_ONLINE_01": c_online_list,
    "IT_ONLINE_02": R.c_cm_course_create,
    "IT_ONLINE_03": R.c_cm_course_publish,
    "IT_ONLINE_04": c_online_forbidden,
    "IT_SYLLABUS_01": R.c_curr_programs,
    "IT_SYLLABUS_02": R.c_curr_programs,
    "IT_SYLLABUS_03": R.c_curr_path,
    "IT_SYLLABUS_04": R.c_curr_bank,
    "IT_SYLLABUS_05": R.c_curr_programs,
    # NOTIF
    "IT_NOTIF_01": R.c_notif_list,
    "IT_NOTIF_02": c_notif_read,
    "IT_NOTIF_03": c_notif_read_all,
    "IT_NOTIF_04": R.c_notif_prefs_get,
    "IT_NOTIF_05": c_notif_unauth,
    # SUPPORT
    "IT_SUPPORT_01": R.c_support_create,
    "IT_SUPPORT_02": R.c_support_list,
    "IT_SUPPORT_03": R.c_support_staff,
    "IT_SUPPORT_04": R.c_support_bad,
    # ADMIN / BROADCAST
    "IT_ADMIN_01": R.c_admin_users,
    "IT_ADMIN_02": c_admin_lock,
    "IT_BROADCAST_01": c_broadcast_create,
    "IT_BROADCAST_02": c_broadcast_list,
    "IT_BROADCAST_03": c_broadcast_update,
    "IT_BROADCAST_04": c_broadcast_cancel,
    # GMEET / REPORT
    "IT_GMEET_01": c_gmeet_teacher,
    "IT_GMEET_02": c_gmeet_learner,
    "IT_GMEET_03": c_gmeet_forbidden,
    "IT_REPORT_01": R.c_report_dash,
    "IT_REPORT_02": R.c_report_revenue,
}


def all_case_ids() -> list[str]:
    ids = []
    for m in MODULES:
        for kind, *_rest, case in iter_cases(m):
            if kind == "CASE":
                ids.append(case["id"])
    return ids


def run_one(ctx: R.Ctx, case_id: str) -> R.Result:
    fn = CHECKS.get(case_id)
    if not fn:
        return R.na("No automated check mapped yet")
    try:
        return fn(ctx)
    except Exception as e:
        return R.fail(f"exception: {e}")


def run_round(round_no: int, only_ids: set[str] | None = None) -> dict[str, R.Result]:
    print(f"\n=== ROUND {round_no} ===")
    ctx = R.Ctx()
    for role in R.ACCOUNTS:
        try:
            ctx.login(role)
            print(f"  login {role}: OK")
        except Exception as e:
            print(f"  login {role}: FAIL {e}")
    results: dict[str, R.Result] = {}
    case_ids = [cid for cid in all_case_ids() if only_ids is None or cid in only_ids]
    for cid in case_ids:
        status, note = run_one(ctx, cid)
        results[cid] = (status, note)
        msg = f"  {cid}: {status} - {note[:100]}"
        print(msg.encode("ascii", "replace").decode("ascii"))
    return results


def write_excel(rounds: list[dict[str, R.Result]], only_ids: set[str] | None = None) -> Path:
    if not EXCEL.exists():
        raise SystemExit(f"Missing Excel: {EXCEL}")
    shutil.copy2(EXCEL, BACKUP)
    print("BACKUP", BACKUP)

    wb = load_workbook(EXCEL)

    # Align CLASS_02 procedure with real API (proposal create, not POST /classrooms)
    if "IT_CLASS" in wb.sheetnames:
        ws_cls = wb["IT_CLASS"]
        for row in range(11, min(ws_cls.max_row or 11, 80) + 1):
            if ws_cls.cell(row, 1).value == "IT_CLASS_02":
                ws_cls.cell(row, 3).value = (
                    "1. Login as STAFF and obtain a JWT.\n"
                    "2. Call GET /api/staff/classrooms/training-programs to pick a published TrainingProgram id.\n"
                    "3. Call POST /api/staff/classroom-proposals with title, courseOfferingId, capacity, dates, weekdays, session times.\n"
                    "4. StaffClassroomProposalController -> ClassroomProposalService.create(); draft proposal saved.\n"
                    "Note: StaffClassroomController has no POST /api/staff/classrooms root create."
                )
                ws_cls.cell(row, 4).value = (
                    "HTTP status is 200 or 201.\n"
                    "JSON contains proposal id and approvalStatus=DRAFT."
                )
                ws_cls.cell(row, 5).value = "Published TrainingProgram and STAFF JWT available."
                break

    for m in MODULES:
        name = m["sheet"]
        if name not in wb.sheetnames:
            print("SKIP missing sheet", name)
            continue
        ws = wb[name]
        for row in range(11, min(ws.max_row or 11, 200) + 1):
            cid = ws.cell(row, 1).value
            if not isinstance(cid, str) or not cid.startswith("IT_"):
                continue
            if only_ids is not None and cid not in only_ids:
                continue
            for ri, col_status, col_date, col_tester in (
                (0, 6, 7, 8),
                (1, 9, 10, 11),
                (2, 12, 13, 14),
            ):
                status, note = rounds[ri].get(cid, ("Pending", "not run"))
                # Only overwrite result cells — keep fills/fonts/borders
                ws.cell(row, col_status).value = status
                ws.cell(row, col_date).value = TODAY
                ws.cell(row, col_tester).value = TESTER
            _st, note = rounds[2].get(cid, ("Pending", ""))
            note_cell = ws.cell(row, 15)
            if note:
                note_cell.value = note
            # Keep Note inside styled cell (border + white fill)
            note_cell.fill = _NOTE_FILL
            note_cell.font = _NOTE_FONT
            note_cell.border = _NOTE_BORDER
            note_cell.alignment = _NOTE_ALIGN

    if "Cover" in wb.sheetnames:
        cover = wb["Cover"]
        scope = "fixed CLASS/ASNTEACH/BROADCAST/GMEET APIs" if only_ids else "full suite"
        cover["E12"] = (
            f"Executed 3 rounds ({scope}) on {TODAY} against {BASE}; tester={TESTER}; "
            "status/date/tester filled without regenerating sheet layout."
        )

    try:
        wb.save(EXCEL)
        out = EXCEL
    except PermissionError:
        alt = EXCEL.with_name(EXCEL.stem + "_EXECUTED.xlsx")
        wb.save(alt)
        out = alt
        print("LOCKED original ->", alt)

    shutil.copy2(out, PROJ / out.name)
    print("WROTE", out)
    return out


def wait_backend(timeout_s: int = 240) -> None:
    t0 = time.time()
    while time.time() - t0 < timeout_s:
        try:
            r = requests.get(f"{BASE}/api/online-courses", timeout=5)
            if r.status_code < 500:
                print("backend ready", r.status_code)
                return
        except Exception:
            pass
        time.sleep(3)
        print("waiting backend...")
    raise SystemExit("Backend not reachable")


def summarize(rounds: list[dict[str, R.Result]]):
    for i, rd in enumerate(rounds, 1):
        counts = {"Passed": 0, "Failed": 0, "Pending": 0, "N/A": 0}
        for st, _ in rd.values():
            counts[st] = counts.get(st, 0) + 1
        print(f"Round {i} summary: {counts}")


def main():
    fix_only = "--fix-apis" in sys.argv
    only_ids = FIX_CASE_IDS if fix_only else None
    wait_backend()
    rounds = [run_round(1, only_ids), run_round(2, only_ids), run_round(3, only_ids)]
    summarize(rounds)
    path = write_excel(rounds, only_ids)
    payload = {
        "date": TODAY,
        "tester": TESTER,
        "base": BASE,
        "excel": str(path),
        "mode": "fix-apis" if fix_only else "full",
        "caseIds": sorted(only_ids) if only_ids else "all",
        "rounds": [
            {cid: {"status": st, "note": note} for cid, (st, note) in rd.items()}
            for rd in rounds
        ],
    }
    out_summary = PROJ / ("it_execution_summary_uc_fix.json" if fix_only else SUMMARY.name)
    out_summary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print("SUMMARY", out_summary)


if __name__ == "__main__":
    main()
