# -*- coding: utf-8 -*-
from pathlib import Path

from openpyxl import load_workbook

p = Path(r"C:\Users\phong\Downloads\intergration test\SEP490_G23_Report5.2_Integration Test_COMPLETED_v4_FORMATTED.xlsx")
wb = load_workbook(p)
sheets = set(wb.sheetnames)

tc = wb["Test Cases"]
print("TC D3", tc["D3"].value, "D4", tc["D4"].value)
bad = []
for r in range(9, 40):
    cell = tc.cell(r, 4)
    if cell.value is None:
        break
    loc = cell.hyperlink.location if cell.hyperlink else None
    expected = f"'{cell.value}'!A1"
    ok = loc == expected and cell.value in sheets
    print(f"  TC R{r} display={cell.value!r} loc={loc!r} ok={ok}")
    if not ok:
        bad.append(("TC", r, cell.value, loc))

st = wb["Test Statistics"]
print("ST C3", st["C3"].value, "C4", st["C4"].value)
for r in range(11, 50):
    cell = st.cell(r, 3)
    if cell.value is None:
        continue
    if str(cell.value).strip().lower() == "sub total":
        print("  SUBTOTAL row", r, [st.cell(r, c).value for c in range(4, 9)])
        print("  Coverage", st.cell(r + 2, 3).value, st.cell(r + 2, 5).value)
        print("  Success", st.cell(r + 3, 3).value, st.cell(r + 3, 5).value)
        break
    loc = cell.hyperlink.location if cell.hyperlink else None
    sheet = None
    if loc and loc.startswith("'") and loc.endswith("'!A1"):
        sheet = loc[1:-4]
    ok = sheet in sheets and cell.hyperlink is not None
    print(f"  ST R{r} display={cell.value!r} loc={loc!r} sheet_ok={ok}")
    if not ok:
        bad.append(("ST", r, cell.value, loc))

print("Auth B2", wb["IT - Auth"]["B2"].value)
print("BAD", bad or "None")
print("cases", sum(
    1
    for s in wb.sheetnames
    if s.startswith("IT - ")
    for row in wb[s].iter_rows(min_col=1, max_col=1, values_only=True)
    if row[0] and str(row[0]).startswith("IT_")
))
