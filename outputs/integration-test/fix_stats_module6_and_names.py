# -*- coding: utf-8 -*-
"""Fix Test Statistics module-6 percent formats; set Creator/Tester to phongdx."""
from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, numbers

ROOT = Path(r"C:\Users\phong\Downloads\intergration test")
SRC = ROOT / "SEP490_G23_Report5.2_Integration Test_COMPLETED_v4_FORMATTED.xlsx"
OUT = ROOT / "SEP490_G23_Report5.2_Integration Test_COMPLETED_v4_FORMATTED.xlsx"
OUT_EXEC = ROOT / "SEP490_G23_Report5.2_Integration Test_COMPLETED_v4_EXECUTED.xlsx"
PROJ = Path(r"D:\EngLishLab\EnglishLab\outputs\integration-test")

CREATOR = "phongdx"
TESTER = "phongdx"
VALUE_FONT = Font(name="Tahoma", size=10, color="008000")

wb = load_workbook(SRC)

# --- Cover + Statistics creator ---
cover = wb["Cover"]
cover["F4"] = CREATOR

st = wb["Test Statistics"]
st["F3"] = CREATOR
st["F3"].font = VALUE_FONT

# Fix number formats on all module rows + subtotal (not coverage %)
# Module rows 11..34, subtotal 36
for r in range(11, 37):
    for c in range(2, 9):
        cell = st.cell(r, c)
        # Keep module code / labels as general; force counts to integer general
        if c == 2 or c == 3:
            cell.number_format = "General"
        else:
            # count columns
            cell.number_format = "0"

# Coverage rows stay percentage-like display via formula returning 0-100 + "%" label in next col
for r in (38, 39):
    cell = st.cell(r, 5)
    cell.number_format = "0.00"

# --- Tester on all IT sheets ---
for name in wb.sheetnames:
    if not name.startswith("IT - "):
        continue
    ws = wb[name]
    for r in range(11, ws.max_row + 1):
        cid = ws.cell(r, 1).value
        if not (isinstance(cid, str) and cid.startswith("IT_")):
            continue
        for col in (8, 11, 14):  # Tester Round1/2/3
            if ws.cell(r, col).value not in (None, ""):
                ws.cell(r, col).value = TESTER

wb.save(OUT)
shutil.copy2(OUT, PROJ / OUT.name)
# keep EXECUTED in sync if present
if OUT_EXEC.exists():
    wb2 = load_workbook(OUT_EXEC)
    wb2["Cover"]["F4"] = CREATOR
    wb2["Test Statistics"]["F3"] = CREATOR
    for r in range(11, 37):
        for c in range(2, 9):
            cell = wb2["Test Statistics"].cell(r, c)
            cell.number_format = "General" if c in (2, 3) else "0"
    for name in wb2.sheetnames:
        if not name.startswith("IT - "):
            continue
        ws = wb2[name]
        for r in range(11, ws.max_row + 1):
            cid = ws.cell(r, 1).value
            if isinstance(cid, str) and cid.startswith("IT_"):
                for col in (8, 11, 14):
                    if ws.cell(r, col).value not in (None, ""):
                        ws.cell(r, col).value = TESTER
    wb2.save(OUT_EXEC)
    shutil.copy2(OUT_EXEC, PROJ / OUT_EXEC.name)

print("fixed", OUT)
print("creator", CREATOR, "tester", TESTER)
