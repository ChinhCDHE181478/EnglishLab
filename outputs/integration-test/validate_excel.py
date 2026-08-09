from openpyxl import load_workbook
from pathlib import Path

p = Path(r"C:\Users\phong\Downloads\intergration test\SEP490_G23_Report5.2_Integration Test_COMPLETED.xlsx")
wb = load_workbook(p)
print("SHEETS:", wb.sheetnames)
assert "Feature 1" not in wb.sheetnames and "Feature 2" not in wb.sheetnames

ids = []
for name in wb.sheetnames:
    if not name.startswith("IT -"):
        continue
    ws = wb[name]
    print("---", name, "B2=", ws["B2"].value, "B4=", ws["B4"].value)
    for row in ws.iter_rows(min_row=11, max_col=6, values_only=True):
        a = row[0]
        if isinstance(a, str) and a.startswith("IT_"):
            ids.append(a)
            assert row[5] == "Pending", (a, row[5])
            assert row[1] and row[2] and row[3] and row[4]

print("CASE COUNT", len(ids), "UNIQUE", len(set(ids)))
assert len(ids) == len(set(ids))

tc = wb["Test Cases"]
print("TC project", tc["D3"].value, tc["D4"].value)
for r in range(9, 25):
    sheet = tc.cell(r, 4).value
    if sheet:
        print(" TC row", r, tc.cell(r, 3).value, "->", sheet, "exists", sheet in wb.sheetnames)
        assert sheet in wb.sheetnames

st = wb["Test Statistics"]
print("STAT C3", st["C3"].value)
for r in range(11, 30):
    c = st.cell(r, 3).value
    if c:
        print(" STAT", r, c, st.cell(r, 4).value, st.cell(r, 8).value)
        for col in range(3, 9):
            v = st.cell(r, col).value
            if isinstance(v, str) and "#REF!" in v:
                raise SystemExit(f"REF at {r},{col}")

refs = []
for name in wb.sheetnames:
    ws = wb[name]
    for row in ws.iter_rows(max_row=min(ws.max_row, 250), max_col=min(ws.max_column or 20, 20)):
        for cell in row:
            if isinstance(cell.value, str) and "#REF!" in cell.value:
                refs.append((name, cell.coordinate, cell.value))
print("REFS", refs or "None")
print("OK")
