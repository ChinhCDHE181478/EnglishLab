# -*- coding: utf-8 -*-
"""
Split IT_SCHEDULE into IT_SCHEDULE / IT_ATTEND / IT_MNGHW,
preserve Round results, refresh Test Cases + Test Statistics.
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parent))

import rebuild_it_excel_from_uc as reb  # noqa: E402
from uc_modules import MODULES, iter_cases  # noqa: E402

EXCEL = Path(
    r"C:\Users\phong\Downloads\intergration test"
    r"\SEP490_G23_Report5.2_Integration Test_COMPLETED.xlsx"
)
PROJ = Path(r"D:\EngLishLab\EnglishLab\outputs\integration-test")
BACKUP = EXCEL.with_name(EXCEL.stem + "_BEFORE_SPLIT_SCHEDULE.xlsx")

# Old IT_SCHEDULE_* -> new IDs
ID_MAP = {
    "IT_SCHEDULE_01": "IT_SCHEDULE_01",
    "IT_SCHEDULE_02": "IT_SCHEDULE_02",
    "IT_SCHEDULE_03": "IT_ATTEND_01",
    "IT_SCHEDULE_04": "IT_ATTEND_02",
    "IT_SCHEDULE_05": "IT_MNGHW_01",
    "IT_SCHEDULE_06": "IT_MNGHW_02",
}

GROUP_FILL = PatternFill("solid", fgColor="CCFFFF")
GROUP_FONT = Font(name="Tahoma", size=10, bold=True, color="000000")
CASE_FILL = PatternFill("solid", fgColor="FFFFFF")
NORMAL_FONT = Font(name="Tahoma", size=10, color="000000")
HEADER_FILL = PatternFill("solid", fgColor="76923C")
HEADER_FONT = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")
THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

STATUS_COLS = (6, 9, 12)
DATE_COLS = (7, 10, 13)
TESTER_COLS = (8, 11, 14)
IT_HEADERS = [
    "Test Case ID",
    "Test Case Description",
    "Test Case Procedure",
    "Expected Results",
    "Pre-conditions",
    "Round 1",
    "Test date",
    "Tester",
    "Round 2",
    "Test date",
    "Tester",
    "Round 3",
    "Test date",
    "Tester",
    "Note",
]


def read_results(ws) -> dict[str, dict]:
    out: dict[str, dict] = {}
    for row in range(11, min(ws.max_row or 11, 300) + 1):
        cid = ws.cell(row, 1).value
        if not isinstance(cid, str) or not cid.startswith("IT_"):
            continue
        out[cid] = {
            "status": [ws.cell(row, c).value for c in STATUS_COLS],
            "date": [ws.cell(row, c).value for c in DATE_COLS],
            "tester": [ws.cell(row, c).value for c in TESTER_COLS],
            "note": ws.cell(row, 15).value,
        }
    return out


def ensure_round_formulas(ws):
    ws["A2"] = "Feature"
    ws["A3"] = "Test requirement"
    ws["A4"] = "Number of TCs"
    ws["B4"] = '=COUNTIF(A11:A1000,"IT_*")'
    ws["A5"] = "Testing Round"
    ws["B5"] = "Passed"
    ws["C5"] = "Failed"
    ws["D5"] = "Pending"
    ws["E5"] = "N/A"
    # Round count formulas (same pattern as other IT sheets)
    for r, cols in (
        (6, ("$F11:$F991",)),
        (7, ("$I11:$I991",)),
        (8, ("$L11:$L991",)),
    ):
        pass
    ws["A6"] = "Round 1"
    ws["A7"] = "Round 2"
    ws["A8"] = "Round 3"
    for col, letter in ((2, "B"), (3, "C"), (4, "D"), (5, "E")):
        label = ws.cell(5, col).value  # Passed/Failed/...
        ws.cell(6, col).value = f'=COUNTIF($F11:$F991,{letter}5)'
        ws.cell(7, col).value = f'=COUNTIF($I11:$I991,{letter}5)'
        ws.cell(8, col).value = f'=COUNTIF($L11:$L991,{letter}5)'

    for c, label in enumerate(IT_HEADERS, start=1):
        cell = ws.cell(10, c)
        cell.value = label
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
        cell.alignment = CENTER


def clear_body(ws, start=11, end=80):
    for r in range(start, end + 1):
        for c in range(1, 16):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.font = NORMAL_FONT
            cell.border = Border()
            cell.alignment = Alignment()
        ws.row_dimensions[r].height = None


def write_module(ws, module, results_by_new_id: dict[str, dict]):
    ensure_round_formulas(ws)
    ws["B2"] = module["function"]
    ws["B3"] = module["requirement"]
    clear_body(ws)
    row = 11
    for kind, gname, _scope, case in iter_cases(module):
        if kind == "GROUP":
            ws.cell(row, 1).value = gname
            for c in range(1, 16):
                cell = ws.cell(row, c)
                cell.fill = GROUP_FILL
                cell.font = GROUP_FONT
                cell.border = THIN
                cell.alignment = Alignment(vertical="center")
                if c > 1:
                    cell.value = None
            ws.row_dimensions[row].height = 18
            row += 1
            continue

        ws.cell(row, 1).value = case["id"]
        ws.cell(row, 2).value = case["desc"]
        ws.cell(row, 3).value = case["proc"]
        ws.cell(row, 4).value = case["exp"]
        ws.cell(row, 5).value = case["pre"]
        for c in range(1, 16):
            cell = ws.cell(row, c)
            cell.fill = CASE_FILL
            cell.font = NORMAL_FONT
            cell.border = THIN
            cell.alignment = WRAP if c <= 5 or c == 15 else CENTER

        prev = results_by_new_id.get(case["id"], {})
        statuses = prev.get("status") or ["Pending", "Pending", "Pending"]
        dates = prev.get("date") or [None, None, None]
        testers = prev.get("tester") or [None, None, None]
        for i, c in enumerate(STATUS_COLS):
            ws.cell(row, c).value = statuses[i] if i < len(statuses) else "Pending"
        for i, c in enumerate(DATE_COLS):
            ws.cell(row, c).value = dates[i] if i < len(dates) else None
        for i, c in enumerate(TESTER_COLS):
            ws.cell(row, c).value = testers[i] if i < len(testers) else None
        ws.cell(row, 15).value = prev.get("note")
        ws.row_dimensions[row].height = 60
        row += 1


def clone_sheet(wb, template: str, target: str):
    if target in wb.sheetnames:
        return wb[target]
    ws = wb.copy_worksheet(wb[template])
    ws.title = target
    print("CLONED", template, "->", target)
    return ws


def main():
    if not EXCEL.exists():
        raise SystemExit(f"Missing {EXCEL}")
    shutil.copy2(EXCEL, BACKUP)
    wb = load_workbook(EXCEL)

    if "IT_SCHEDULE" not in wb.sheetnames:
        raise SystemExit("Missing IT_SCHEDULE")

    old = read_results(wb["IT_SCHEDULE"])
    # Also pull from ATTEND/MNGHW if already partially present
    for name in ("IT_ATTEND", "IT_MNGHW", "IT_TCHHW"):
        if name in wb.sheetnames:
            old.update(read_results(wb[name]))

    results_new: dict[str, dict] = {}
    for old_id, new_id in ID_MAP.items():
        if old_id in old:
            results_new[new_id] = old[old_id]
        elif new_id in old:
            results_new[new_id] = old[new_id]
    # migrate old TCHHW ids if present
    for a, b in (("IT_TCHHW_01", "IT_MNGHW_01"), ("IT_TCHHW_02", "IT_MNGHW_02")):
        if a in old and b not in results_new:
            results_new[b] = old[a]

    clone_sheet(wb, "IT_SCHEDULE", "IT_ATTEND")
    clone_sheet(wb, "IT_SCHEDULE", "IT_MNGHW")
    if "IT_TCHHW" in wb.sheetnames and "IT_MNGHW" in wb.sheetnames:
        wb.remove(wb["IT_TCHHW"])

    by_sheet = {m["sheet"]: m for m in MODULES}
    for sheet in ("IT_SCHEDULE", "IT_ATTEND", "IT_MNGHW"):
        write_module(wb[sheet], by_sheet[sheet], results_new)
        print("WROTE", sheet, by_sheet[sheet]["function"])

    reb.rewrite_test_cases_sheet(wb["Test Cases"])
    reb.rewrite_statistics(wb["Test Statistics"])
    reb.reorder_it_sheets(wb)

    # Drop obsolete sheets not in MODULES
    keep = {"Cover", "Test Cases", "Test Statistics"} | {m["sheet"] for m in MODULES}
    for name in list(wb.sheetnames):
        if name not in keep:
            wb.remove(wb[name])
            print("REMOVED", name)

    try:
        wb.save(EXCEL)
        out = EXCEL
    except PermissionError:
        out = EXCEL.with_name(EXCEL.stem + "_SPLIT_SCHEDULE.xlsx")
        wb.save(out)
        print("LOCKED ->", out)

    shutil.copy2(out, PROJ / out.name)
    print("BACKUP", BACKUP)
    print("WROTE", out)
    print("modules", len(MODULES))


if __name__ == "__main__":
    main()
