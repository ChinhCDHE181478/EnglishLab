# -*- coding: utf-8 -*-
"""Style Note column (O) on all IT_* sheets: border + fill like other result cells."""
from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

EXCEL = Path(
    r"C:\Users\phong\Downloads\intergration test"
    r"\SEP490_G23_Report5.2_Integration Test_COMPLETED.xlsx"
)
PROJ = Path(r"D:\EngLishLab\EnglishLab\outputs\integration-test")
BACKUP = EXCEL.with_name(EXCEL.stem + "_BEFORE_NOTE_STYLE.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="76923C")
HEADER_FONT = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")
GROUP_FILL = PatternFill("solid", fgColor="CCFFFF")
CASE_FILL = PatternFill("solid", fgColor="FFFFFF")
NORMAL_FONT = Font(name="Tahoma", size=10, color="000000")
THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def main() -> None:
    if not EXCEL.exists():
        raise SystemExit(f"Missing {EXCEL}")
    shutil.copy2(EXCEL, BACKUP)
    wb = load_workbook(EXCEL)
    n = 0
    for name in wb.sheetnames:
        if not name.startswith("IT_"):
            continue
        ws = wb[name]
        # Header Note
        h = ws.cell(10, 15)
        h.value = "Note"
        h.fill = HEADER_FILL
        h.font = HEADER_FONT
        h.border = THIN
        h.alignment = CENTER
        # Body rows
        for row in range(11, min(ws.max_row or 11, 300) + 1):
            a = ws.cell(row, 1).value
            if a is None and ws.cell(row, 15).value is None:
                continue
            cell = ws.cell(row, 15)
            is_group = isinstance(a, str) and not a.startswith("IT_")
            is_case = isinstance(a, str) and a.startswith("IT_")
            if not (is_group or is_case or cell.value is not None):
                continue
            cell.fill = GROUP_FILL if is_group else CASE_FILL
            cell.font = Font(name="Tahoma", size=10, bold=is_group)
            cell.border = THIN
            cell.alignment = Alignment(vertical="center") if is_group else WRAP
            n += 1
        # Reasonable Note width
        ws.column_dimensions[get_column_letter(15)].width = max(
            ws.column_dimensions[get_column_letter(15)].width or 0, 28
        )

    try:
        wb.save(EXCEL)
        out = EXCEL
    except PermissionError:
        out = EXCEL.with_name(EXCEL.stem + "_NOTE_STYLED.xlsx")
        wb.save(out)
        print("LOCKED original ->", out)
    shutil.copy2(out, PROJ / out.name)
    print("styled cells", n, "BACKUP", BACKUP, "WROTE", out)


if __name__ == "__main__":
    main()
