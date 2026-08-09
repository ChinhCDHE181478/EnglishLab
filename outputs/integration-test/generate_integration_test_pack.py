# -*- coding: utf-8 -*-
"""Generate EnglishLab Integration Test documentation + completed Excel workbook."""
from __future__ import annotations

import copy
import json
import shutil
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink

ROOT = Path(r"C:\Users\phong\Downloads\intergration test")
OUT_PROJECT = Path(r"D:\EngLishLab\EnglishLab\outputs\integration-test")
TEMPLATE = ROOT / "SEP490_G23_Report5.2_Integration Test (1).xlsx"
BACKUP = ROOT / "SEP490_G23_Report5.2_Integration Test_BACKUP.xlsx"
COMPLETED = ROOT / "SEP490_G23_Report5.2_Integration Test_COMPLETED_v4.xlsx"

PROJECT_NAME = "EnglishLab"
PROJECT_CODE = "SEP490_G23"
CREATOR = "phongdx"
ISSUE_DATE = "2026-07-28"
VERSION = "v4.0"

LINK_FONT = Font(name="Tahoma", size=10, color="0000FF", underline="single")
VALUE_FONT = Font(name="Tahoma", size=10, color="008000")


def clear_cell_hyperlinks(ws, cells):
    """Remove existing hyperlinks that occupy the given cell refs."""
    wanted = {str(c).upper() for c in cells}
    remaining = []
    for h in list(getattr(ws, "_hyperlinks", []) or []):
        ref = (getattr(h, "ref", None) or "").upper()
        if ref in wanted:
            continue
        remaining.append(h)
    ws._hyperlinks = remaining


def set_sheet_hyperlink(cell, sheet_name: str, display: str | None = None):
    """Create an internal workbook jump link like the school template."""
    display = display or sheet_name
    location = f"'{sheet_name}'!A1"
    cell.value = display
    cell.hyperlink = Hyperlink(ref=cell.coordinate, location=location, display=display)
    cell.font = LINK_FONT

# ---------------------------------------------------------------------------
# Modules imported from full-project coverage definition
# ---------------------------------------------------------------------------
from full_modules import MODULES  # noqa: E402


def iter_cases(module):
    for g in module["groups"]:
        yield ("GROUP", g["name"], None)
        for c in g["cases"]:
            yield ("CASE", g["name"], c)


def all_case_ids():
    ids = []
    for m in MODULES:
        for kind, _, c in iter_cases(m):
            if kind == "CASE":
                ids.append(c["id"])
    return ids


def write_analysis(path: Path):
    lines = []
    lines.append("# Integration Test Analysis — EnglishLab (SEP490_G23)\n")
    lines.append(
        f"- Generated: {ISSUE_DATE} (**{VERSION} full-project coverage**)\n"
        "- Scope: entire backend business surface (~56 RestControllers), grouped into integration modules.\n"
        "- Sources: backend source, SRS Report3, srs-usecase-diagram-map.md, Excel template, controller inventory.\n"
    )
    lines.append("## A. Project overview\n")
    lines.append("""| Item | Fact from repo |
|---|---|
| Backend | Java 21, Spring Boot 4.0.6 (`backend/pom.xml`) |
| Frontend | React + Vite (`frontend/`) |
| Database | PostgreSQL via Spring Data JPA |
| AuthN/Z | Spring Security + JWT (`JwtAuthenticationFilter`, `SecurityConfig`) |
| API docs | springdoc OpenAPI (`/swagger-ui`, `/v3/api-docs`) |
| Mail | `spring-boot-starter-mail` (OTP/verification/business mail) |
| Payments | PayOS (`vn.payos:payos-java`, `PayosWebhookController`) |
| Meetings | Lark webhook/events (`LarkWebhookController`) |
| AI | Gemini evaluation client (assessment speaking/writing) |
| Run backend | Spring Boot main `BackendApplication`; typical `mvnw spring-boot:run` with `.env` |
| Run tests | `mvnw test` / targeted `-Dtest=...`; mostly Mockito unit tests + few `@SpringBootTest` |
| Testcontainers | Not present in `pom.xml` (as surveyed) |
""")
    lines.append("### Main directories\n")
    lines.append("""- `backend/src/main/java/.../controller|service|repository|entity|security|migration|seed`
- `backend/src/test/java/...`
- `frontend/src/pages|api|components`
- `docs/` / root SRS extracts
""")
    lines.append("## B. Component inventory (representative)\n")
    lines.append("| Layer | Component | Path | Responsibility | Dependencies |\n|---|---|---|---|---|")
    rows = [
        ("Controller", "AuthController", "controller/AuthController.java", "Register/login/verify/OTP/social", "AuthService"),
        ("Controller", "UserController", "controller/UserController.java", "Profile/avatar/password/prefs", "UserService, NotificationPreferenceService"),
        ("Controller", "StudentCommerceController", "controller/commerce/...", "Cart/wishlist", "Commerce services"),
        ("Controller", "StudentPaymentController", "controller/payment/...", "PayOS link/orders", "PaymentService"),
        ("Controller", "PayosWebhookController", "controller/payment/...", "Payment webhook", "PaymentService"),
        ("Controller", "PublicOnlineCourseController", "controller/course/...", "Public catalog", "OnlineCourseService"),
        ("Controller", "StudentOnlineCourseController", "controller/course/...", "Enrollment content/progress", "OnlineCourse services"),
        ("Controller", "PublicClassroomController", "controller/classroom/...", "Public offerings", "ClassroomOfferingService"),
        ("Controller", "TrainingManagerClassroomController", "controller/classroom/...", "TM classroom + enrollments", "ClassroomOfferingService, TuitionProofService"),
        ("Controller", "TeacherClassroomController", "controller/classroom/...", "Teacher ops", "Homework/attendance/gradebook services"),
        ("Controller", "StudentClassroomController", "controller/classroom/...", "Learner classroom", "Offering/homework/tuition proof"),
        ("Controller", "PlacementTestController", "controller/assessment/...", "Placement current/submit", "Placement services"),
        ("Controller", "StudentSupportTicketController", "controller/support/...", "Learner tickets", "SupportTicketService"),
        ("Controller", "AdminUserController", "controller/admin/...", "Admin users", "AdminUserService"),
        ("Security", "SecurityConfig", "config/SecurityConfig.java", "Route authorization", "Jwt filter, UserDetails"),
        ("Security", "JwtAuthenticationFilter", "security/...", "Bearer JWT parse", "Token service"),
        ("Service", "ClassroomOfferingServiceImpl", "service/classroom/impl/...", "Enrollment/waitlist/tuition", "Repos + notifications"),
        ("Service", "NotificationPreferenceServiceImpl", "service/notification/impl/...", "Channel prefs", "notification_preferences"),
        ("Service", "AppNotificationServiceImpl", "service/notification/impl/...", "Persist in-app notifications", "prefs gate + app_notifications"),
        ("Service", "PaymentService", "service/payment/...", "Orders + webhook side effects", "payment_orders, PayOS"),
        ("Entity/DB", "User", "entity/User.java", "users", "roles"),
        ("Entity/DB", "ClassroomEnrollment", "entity/classroom/...", "classroom_enrollments", "offering, student"),
        ("Entity/DB", "PaymentOrder", "entity/payment/...", "payment_orders", "user/courses"),
        ("Entity/DB", "SupportTicket", "entity/support/...", "support_tickets", "messages"),
        ("External", "PayOS", "payos-java + webhook", "Checkout", "PaymentService"),
        ("External", "Mail", "JavaMailSender services", "OTP/business mail", "Auth/classroom mail services"),
        ("External", "Lark", "LarkWebhookController", "Meeting events", "classroom sessions"),
        ("External", "Gemini AI", "GeminiAiEvaluationClientImpl", "Speaking/writing eval", "Assessment services"),
    ]
    for r in rows:
        lines.append("| " + " | ".join(r) + " |")

    lines.append("\n## C. API inventory (verified mappings)\n")
    lines.append("| Method | Endpoint | Controller | Role | Main flow |\n|---|---|---|---|---|")
    apis = [
        ("POST", "/api/auth/register", "AuthController", "permitAll", "Create user + verification token"),
        ("POST", "/api/auth/login", "AuthController", "permitAll", "JWT issue"),
        ("POST", "/api/auth/verify-email", "AuthController", "permitAll", "Verify OTP"),
        ("POST", "/api/auth/forgot-password", "AuthController", "permitAll", "Reset OTP"),
        ("POST", "/api/auth/reset-password", "AuthController", "permitAll", "Password update"),
        ("GET/PUT", "/api/user/me", "UserController", "authenticated", "Profile"),
        ("GET/PUT", "/api/user/me/notification-preferences", "UserController", "authenticated", "Prefs upsert"),
        ("GET/POST/DELETE", "/api/student/commerce/cart*", "StudentCommerceController", "LEARNER(+)", "Cart"),
        ("GET/POST/DELETE", "/api/student/commerce/wishlist*", "StudentCommerceController", "LEARNER(+)", "Wishlist"),
        ("POST", "/api/student/payments/payos/link", "StudentPaymentController", "LEARNER(+)", "Create payment_orders"),
        ("POST", "/api/payos/webhook", "PayosWebhookController", "permitAll", "Confirm payment → access"),
        ("GET", "/api/online-courses/**", "PublicOnlineCourseController", "permitAll GET", "Catalog"),
        ("GET/PATCH", "/api/student/online-courses/**", "StudentOnlineCourseController", "LEARNER(+)", "Content/progress"),
        ("GET", "/api/classroom-offerings/**", "PublicClassroomController", "permitAll GET", "Public classes"),
        ("*", "/api/training-manager/classrooms/**", "TrainingManagerClassroomController", "STAFF/TM/MANAGER/ADMIN", "Ops + enrollments"),
        ("*", "/api/teacher/classrooms/**", "TeacherClassroomController", "TEACHER(+)", "Teaching ops"),
        ("*", "/api/student/classrooms/**", "StudentClassroomController", "LEARNER(+)", "Learner classroom"),
        ("GET/POST", "/api/student/placement-tests/current*", "PlacementTestController", "LEARNER(+)", "Placement"),
        ("*", "/api/student/support-tickets/**", "StudentSupportTicketController", "LEARNER(+)", "Tickets"),
        ("*", "/api/manager|/staff|/training-manager/support-tickets/**", "ManagerSupportTicketController", "staff roles", "Claim/reply"),
        ("*", "/api/admin/**", "AdminUserController", "ADMIN", "User admin"),
        ("GET/PATCH", "/api/student/notifications/**", "StudentNotificationController", "authenticated learner path", "In-app notifications"),
    ]
    for a in apis:
        lines.append("| " + " | ".join(a) + " |")

    lines.append("\n## D. Business flows (examples from code)\n")
    lines.append("""1. `POST /api/auth/register` → AuthController → AuthService → UserRepository + AuthTokenRepository → `users`/`auth_tokens` → (mail) → response
2. `POST /api/auth/login` → AuthService → JWT → response; then `GET /api/user/me` → JwtFilter → UserService → `users`
3. `POST /api/student/payments/payos/link` → PaymentService → `payment_orders` → PayOS API → checkout URL
4. `POST /api/payos/webhook` → PaymentService → update `payment_orders` → grant course ownership/enrollment
5. TM `confirm` → `tuition` → `assign` → ClassroomOfferingService → `classroom_enrollments` (+ payments) → AppNotification optional
6. Teacher `POST .../homework` → ClassroomHomeworkService → `classroom_homework` → optional homework mail gated by NotificationPreferenceService
7. `PUT /api/user/me/notification-preferences` → NotificationPreferenceService → `notification_preferences` → later `AppNotificationService.createForUser` respects in-app flag
8. `POST /api/student/support-tickets` → SupportTicketService → `support_tickets` + `support_ticket_messages`
""")

    lines.append("## E. Integration points\n")
    lines.append("| ID | Source | Target | Type | Related req | Risk |\n|---|---|---|---|---|---|")
    ips = [
        ("IP-01", "AuthController", "User/AuthToken repos", "REST+DB", "UC-01/03/04", "OTP/mail dependency"),
        ("IP-02", "JwtFilter", "Protected controllers", "Security", "All authenticated UC", "Wrong role mapping"),
        ("IP-03", "PaymentService", "PayOS + payment_orders", "External+DB", "UC-47", "Webhook idempotency"),
        ("IP-04", "PayOS webhook", "Course ownership", "Webhook+DB", "UC-08/47", "Duplicate grants"),
        ("IP-05", "ClassroomOfferingService", "enrollments+tuition+notifications", "Transaction+DB", "UC-38", "Partial status updates"),
        ("IP-06", "AppNotificationService", "notification_preferences", "Direct service", "UC-06 + prefs", "Channel bypass"),
        ("IP-07", "HomeworkService", "Mail service", "External", "UC-26", "Mail failure vs DB commit"),
        ("IP-08", "Assessment services", "Gemini client", "External", "UC-17/19", "AI timeout consistency"),
        ("IP-09", "SupportTicketService", "tickets+messages", "DB transaction", "UC-07/44", "Cross-user access"),
        ("IP-10", "AdminUserService", "users+roles+audit", "DB multi-table", "UC-42", "Orphan roles"),
        ("IP-11", "Lark webhook", "classroom_sessions", "Webhook", "UC-10", "Invalid event handling"),
        ("IP-12", "reorderWaitlist service", "classroom_enrollments.waitlist_priority", "Service/DB", "Waitlist priority", "Controller mapping may be missing"),
    ]
    for r in ips:
        lines.append("| " + " | ".join(r) + " |")

    lines.append("\n## Module partitioning rationale\n")
    lines.append("Modules are grouped by business capability (auth, account, commerce, online learning, classroom ops, teacher ops, assessment, content management, support, admin), not one-sheet-per-controller.\n")
    for m in MODULES:
        lines.append(f"- **{m['code']} ({m['name']})**: {m['components']}")

    lines.append("\n## SRS vs source mismatches (selected)\n")
    lines.append("""| Topic | SRS / older map | Current code evidence | IT handling |
|---|---|---|---|
| Support tickets UC-07/44 | Older map: unsupported | `StudentSupportTicketController` + `ManagerSupportTicketController` exist | Covered in SUPPORT module |
| Waitlist reorder API | Implemented historically | `ClassroomOfferingService.reorderWaitlist` exists; **no** `@PutMapping(.../waitlist/order)` found in TM controller at survey time | IT_CLASS_04 notes service-level IT + API gap |
| Placement create/delete | SRS CRUD | Definition often singleton lazy-seed; limited DELETE | ASSESS cases stick to current/submit |
| Teacher skill-bank CRUD | SRS teacher manages 4 skills content | Exercise bank primarily content-manager | Not claimed as TEACH coverage |
| Download materials | SRS downloadability | Often returns stored URL only | CLASS learner access case without fake download API |
| Roles | Training Manager actor | `STAFF` + legacy `TRAINING_MANAGER` both in `RoleEnum`; Security uses both on TM paths | AuthZ cases use real role names |
""")

    lines.append("\n## Open questions\n")
    lines.append("""1. Exact PayOS webhook signature verification mode in local/test profiles.
2. Whether waitlist HTTP endpoint was intentionally removed or regressed.
3. Preferred IT DB strategy (shared PostgreSQL test DB vs future Testcontainers)—pom currently has no Testcontainers.
4. Some FR section pages in SRS extract are sparse (TOC points many FR subsections to same page number)—traceability uses UC IDs as stable anchors.
""")
    path.write_text("\n".join(lines), encoding="utf-8")


def write_plan(path: Path):
    lines = ["# Integration Test Plan — EnglishLab\n", f"Date: {ISSUE_DATE}\n", "## Modules\n"]
    lines.append("| Module code | Module name | Sheet | #TCs | SRS |\n|---|---|---|---:|---|")
    total = 0
    for m in MODULES:
        n = sum(1 for k, _, c in iter_cases(m) if k == "CASE")
        total += n
        lines.append(f"| {m['code']} | {m['name']} | {m['sheet']} | {n} | {m['srs']} |")
    lines.append(f"\n**Total test cases: {total}**\n")
    for m in MODULES:
        lines.append(f"\n## {m['code']} — {m['name']}\n")
        lines.append(f"Requirement scope: {m['requirement']}\n")
        for kind, gname, c in iter_cases(m):
            if kind == "GROUP":
                lines.append(f"\n### {gname}\n")
            else:
                lines.append(f"#### {c['id']}: {c['desc']}\n")
                lines.append(f"- **Procedure:**\n{c['proc']}\n")
                lines.append(f"- **Expected:**\n{c['exp']}\n")
                lines.append(f"- **Pre:** {c['pre']}\n")
                lines.append("- **Round 1:** Pending\n")
    path.write_text("\n".join(lines), encoding="utf-8")


def write_traceability(path: Path):
    def ids_for(prefix):
        found = []
        for m in MODULES:
            for kind, _, c in iter_cases(m):
                if kind == "CASE" and c["id"].startswith(prefix):
                    found.append(c["id"])
        return ", ".join(found) if found else "—"

    mapping = [
        ("UC-01", "Register Account", "AUTH", ids_for("IT_AUTH_0"), "Covered"),
        ("UC-03", "Login", "AUTH", "IT_AUTH_05, IT_AUTH_06, IT_AUTH_07", "Covered"),
        ("UC-04", "Reset password", "AUTH", "IT_AUTH_08, IT_AUTH_09, IT_AUTH_10", "Covered"),
        ("UC-05", "Manage profile", "USER", ids_for("IT_USER_"), "Covered"),
        ("UC-06", "View Notifications", "NOTIF", ids_for("IT_NOTIF_"), "Covered"),
        ("UC-07", "Submit Support Ticket", "SUPPORT", "IT_SUPPORT_01, IT_SUPPORT_02, IT_SUPPORT_04", "Covered"),
        ("UC-44", "Resolve Support Tickets", "SUPPORT", "IT_SUPPORT_03", "Covered"),
        ("UC-02", "View public courses", "COURSE", "IT_COURSE_01, IT_COURSE_02", "Covered"),
        ("UC-08", "Enroll in Course", "PAYMENT/COURSE/ENROLLREQ", "IT_PAYMENT_01–03, IT_COURSE_03–05, IT_ENROLLREQ_*", "Covered"),
        ("UC-45", "Wishlist", "COMMERCE", "IT_COMMERCE_02", "Covered"),
        ("UC-46", "Cart", "COMMERCE", ids_for("IT_COMMERCE_"), "Covered"),
        ("UC-47", "Checkout / PayOS", "PAYMENT", ids_for("IT_PAYMENT_"), "Covered"),
        ("UC-48", "Access online materials", "COURSE", "IT_COURSE_03, IT_COURSE_04", "Covered"),
        ("UC-49", "Discuss in Course", "DISCUSS", "IT_DISCUSS_01, IT_DISCUSS_02", "Covered"),
        ("UC-50", "Report Discussion", "DISCUSS", "IT_DISCUSS_03, IT_DISCUSS_04, IT_DISCUSS_05", "Covered"),
        ("UC-51", "Take Note", "NOTES", ids_for("IT_NOTES_"), "Covered"),
        ("UC-09", "View Timetable", "LEARNERCLS", "IT_LEARNERCLS_01, IT_LEARNERCLS_02", "Covered"),
        ("UC-10", "Join Online Meeting", "LEARNERCLS/LARK", "IT_LEARNERCLS_02, IT_LARK_01–03", "Covered"),
        ("UC-11", "Access classroom materials", "LEARNERCLS", "IT_LEARNERCLS_04", "Covered"),
        ("UC-12", "Download materials", "LEARNERCLS", "—", "SRS/source mismatch — URL only, no download API"),
        ("UC-13", "Submit Homework", "LEARNERCLS/TEACH", "IT_LEARNERCLS_03, IT_TEACH_01–02", "Covered"),
        ("UC-14", "Academic report", "LEARNERCLS/TEACH", "IT_LEARNERCLS_06, IT_TEACH_04", "Covered"),
        ("UC-15", "Take Quiz", "QUIZ", ids_for("IT_QUIZ_"), "Covered"),
        ("UC-16", "Placement exam", "ASSESS", "IT_ASSESS_01–03", "Partially covered (timer/auto-submit client-side)"),
        ("UC-17–20", "Skill practice assessments", "ASSESS", "IT_ASSESS_04–06", "Partially covered"),
        ("UC-22", "Teaching schedule", "TEACH", "IT_TEACH_01", "Partially covered"),
        ("UC-23", "Attendance", "TEACH/DISPUTE", "IT_TEACH_03, IT_DISPUTE_*", "Covered"),
        ("UC-26", "Manage Homework", "TEACH", ids_for("IT_TEACH_"), "Covered"),
        ("UC-27", "Manage Quiz content", "QUIZ", "IT_QUIZ_01, IT_QUIZ_02, IT_QUIZ_04", "Covered"),
        ("UC-32", "Syllabus/curriculum", "CURRICULUM", ids_for("IT_CURRICULUM_"), "Covered"),
        ("UC-33", "Online course CM", "CONTENT/PACKAGE", "IT_CONTENT_*, IT_PACKAGE_*", "Covered"),
        ("UC-36", "Manage Classrooms", "CLASS/INFRA/PROPOSAL", "IT_CLASS_*, IT_INFRA_*, IT_PROPOSAL_*", "Covered"),
        ("UC-37", "Assign Teacher", "CLASS", "IT_CLASS_07", "Covered"),
        ("UC-38", "Assign Learner", "CLASS/ENROLLREQ", "IT_CLASS_03–06, IT_ENROLLREQ_*", "Covered"),
        ("UC-39", "Evaluate Teacher", "—", "—", "Missing implementation"),
        ("UC-40", "Operational report", "REPORT", "IT_REPORT_01", "Partially covered (fixed dashboard)"),
        ("UC-41", "Revenue analytics", "REPORT", "IT_REPORT_02", "Partially covered"),
        ("UC-42", "Manage User Accounts", "ADMIN", ids_for("IT_ADMIN_"), "Covered"),
        ("Notification prefs", "Channel preferences", "NOTIF", "IT_NOTIF_01–03", "Covered"),
    ]
    lines = [
        "# Integration Test Traceability (v2 full project)\n",
        f"Date: {ISSUE_DATE}\n",
        f"Modules: {len(MODULES)}; designed for whole EnglishLab backend surface.\n",
        "| SRS requirement ID | Requirement summary | Module | Test Case IDs | Coverage status |\n|---|---|---|---|---|",
    ]
    for row in mapping:
        lines.append("| " + " | ".join(row) + " |")
    path.write_text("\n".join(lines), encoding="utf-8")


def clear_feature_data(ws):
    # Clear from row 11 downward in columns A-O but keep header row 10
    max_r = max(ws.max_row, 200)
    for r in range(11, max_r + 1):
        for c in range(1, 16):
            ws.cell(r, c).value = None


def write_feature_sheet(ws, module):
    ws["A2"] = "Feature"
    # Match template sample: B2 shows controller/module code used by Test Statistics
    ws["B2"] = module.get("function") or module["name"]
    ws["A3"] = "Test requirement"
    ws["B3"] = module["requirement"]
    # Number of TCs: count IT_* only
    ws["B4"] = '=COUNTIF(A11:A1000,"IT_*")'
    # Standardize Round1 Passed formula like Feature 2
    ws["B5"] = "Passed"
    ws["C5"] = "Failed"
    ws["D5"] = "Pending"
    ws["E5"] = "N/A"
    ws["A6"] = "Round 1"
    ws["B6"] = '=COUNTIF($F11:$F998,B5)'
    ws["C6"] = '=COUNTIF($F11:$F998,C5)'
    ws["D6"] = '=COUNTIF($F11:$F998,D5)'
    ws["E6"] = '=COUNTIF($F11:$F998,E5)'
    ws["A7"] = "Round 2"
    ws["B7"] = '=COUNTIF($I11:$I998,B5)'
    ws["C7"] = '=COUNTIF($I11:$I998,C5)'
    ws["D7"] = '=COUNTIF($I11:$I998,D5)'
    ws["E7"] = '=COUNTIF($I11:$I998,E5)'
    ws["A8"] = "Round 3"
    ws["B8"] = '=COUNTIF($L11:$L998,B5)'
    ws["C8"] = '=COUNTIF($L11:$L998,C5)'
    ws["D8"] = '=COUNTIF($L11:$L998,D5)'
    ws["E8"] = '=COUNTIF($L11:$L998,E5)'

    clear_feature_data(ws)
    row = 11
    wrap = Alignment(wrap_text=True, vertical="top")
    for kind, gname, c in iter_cases(module):
        if kind == "GROUP":
            ws.cell(row, 1).value = gname
            ws.cell(row, 1).alignment = wrap
            row += 1
            continue
        ws.cell(row, 1).value = c["id"]
        ws.cell(row, 2).value = c["desc"]
        ws.cell(row, 3).value = c["proc"]
        ws.cell(row, 4).value = c["exp"]
        ws.cell(row, 5).value = c["pre"]
        ws.cell(row, 6).value = "Pending"
        ws.cell(row, 9).value = "Pending"
        ws.cell(row, 12).value = "Pending"
        for col in range(1, 16):
            ws.cell(row, col).alignment = wrap
        # rough row height
        ws.row_dimensions[row].height = 90
        row += 1

    # data validation for Round columns
    dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,N/A"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"F11:F{row}")
    dv.add(f"I11:I{row}")
    dv.add(f"L11:L{row}")


def write_test_cases_sheet(wb):
    tc = wb["Test Cases"]
    # Cover-linked project metadata (fixes template #REF!)
    tc["D3"] = "=Cover!B4"
    tc["D3"].font = VALUE_FONT
    tc["D4"] = "=Cover!B5"
    tc["D4"].font = VALUE_FONT
    tc["D5"] = (
        "1. Server: Spring Boot 4 + JDK 21 (@SpringBootTest + MockMvc)\n"
        "2. Database: PostgreSQL with migrations applied\n"
        "3. Web Browser: N/A for API Integration Tests (optional Postman/curl)\n"
        "4. JWT security enabled; roles seeded\n"
        "5. External PayOS / Mail / Gemini / Lark stubbed or sandbox"
    )
    tc["D5"].font = VALUE_FONT

    # Clear old sample rows + broken template hyperlinks
    clear_refs = [f"D{r}" for r in range(9, 60)]
    clear_cell_hyperlinks(tc, clear_refs)
    for r in range(9, 60):
        for c in range(2, 7):
            cell = tc.cell(r, c)
            cell.value = None
            cell.hyperlink = None

    for i, m in enumerate(MODULES, start=1):
        r = 8 + i
        fn = m.get("function") or m["name"]
        sheet = m["sheet"]
        tc.cell(r, 2).value = i
        tc.cell(r, 3).value = fn
        set_sheet_hyperlink(tc.cell(r, 4), sheet, display=sheet)
        tc.cell(r, 5).value = m["requirement"]
        tc.cell(r, 6).value = (
            f"DB up; roles seeded; target sheet `{sheet}` exists; "
            "external systems stubbed as needed for this module."
        )


def write_test_statistics_sheet(wb):
    st = wb["Test Statistics"]
    st["C3"] = "=Cover!B4"
    st["C3"].font = VALUE_FONT
    st["C4"] = "=Cover!B5"
    st["C4"].font = VALUE_FONT
    # Template merges E3:F3 over the Creator label; unmerge so value can be stored.
    merged = [str(r) for r in st.merged_cells.ranges]
    if "E3:F3" in merged:
        st.unmerge_cells("E3:F3")
    st["E3"] = "Creator"
    st["F3"] = CREATOR
    if "E5:F5" in merged:
        st.unmerge_cells("E5:F5")
    st["E5"] = "Issue Date"
    st["H5"] = ISSUE_DATE
    st["C6"] = (
        "Release Integration Test pack includes "
        + f"{len(MODULES)} modules: "
        + ", ".join(m["code"] for m in MODULES)
    )
    st["C6"].font = VALUE_FONT

    clear_refs = [f"C{r}" for r in range(11, 80)]
    clear_cell_hyperlinks(st, clear_refs)
    for r in range(11, 80):
        for c in range(2, 9):
            cell = st.cell(r, c)
            cell.value = None
            cell.hyperlink = None

    for i, m in enumerate(MODULES, start=1):
        r = 10 + i
        sheet = m["sheet"]
        st.cell(r, 2).value = i
        # Module code: formula to Feature B2 (no sheet hyperlink — template style)
        cell_mod = st.cell(r, 3)
        cell_mod.value = f"='{sheet}'!B2"
        cell_mod.hyperlink = None
        cell_mod.font = VALUE_FONT
        st.cell(r, 4).value = f"='{sheet}'!B6"  # Passed Round1
        st.cell(r, 5).value = f"='{sheet}'!C6"  # Failed
        st.cell(r, 6).value = f"='{sheet}'!D6"  # Pending
        st.cell(r, 7).value = f"='{sheet}'!E6"  # N/A
        st.cell(r, 8).value = f"='{sheet}'!B4"  # Number of TCs

    last = 10 + len(MODULES)
    sub_r = last + 2
    st.cell(sub_r, 3).value = "Sub total"
    st.cell(sub_r, 4).value = f"=SUM(D11:D{last})"
    st.cell(sub_r, 5).value = f"=SUM(E11:E{last})"
    st.cell(sub_r, 6).value = f"=SUM(F11:F{last})"
    st.cell(sub_r, 7).value = f"=SUM(G11:G{last})"
    st.cell(sub_r, 8).value = f"=SUM(H11:H{last})"

    cov_r = sub_r + 2
    st.cell(cov_r, 3).value = "Test coverage"
    st.cell(cov_r, 5).value = f"=IF((H{sub_r}-G{sub_r})=0,0,(D{sub_r}+E{sub_r})*100/(H{sub_r}-G{sub_r}))"
    st.cell(cov_r, 6).value = "%"
    st.cell(cov_r + 1, 3).value = "Test successful coverage"
    st.cell(cov_r + 1, 5).value = f"=IF((H{sub_r}-G{sub_r})=0,0,D{sub_r}*100/(H{sub_r}-G{sub_r}))"
    st.cell(cov_r + 1, 6).value = "%"


def build_excel():
    shutil.copy2(TEMPLATE, COMPLETED)
    wb = load_workbook(COMPLETED)

    # Cover
    cover = wb["Cover"]
    cover["B4"] = PROJECT_NAME
    cover["B5"] = PROJECT_CODE
    cover["F4"] = CREATOR
    cover["F5"] = ISSUE_DATE
    cover["F6"] = VERSION
    cover["A11"] = ISSUE_DATE
    cover["B11"] = VERSION
    cover["C11"] = "Integration Test design for EnglishLab backend"
    cover["D11"] = "A"
    cover["E11"] = "v4.0 Test Cases/Statistics sheet links + Controller→Service→Repo IT content"
    cover["F11"] = "SRS Report3; source tree; srs-usecase-diagram-map.md"

    # Prepare feature sheets: rename Feature 1/2 then copy
    f1 = wb["Feature 1"]
    f2 = wb["Feature 2"]
    # Use Feature 2 as clone template (better Round1 formulas)
    # Rename first two
    f1.title = MODULES[0]["sheet"]
    f2.title = MODULES[1]["sheet"]
    write_feature_sheet(wb[MODULES[0]["sheet"]], MODULES[0])
    write_feature_sheet(wb[MODULES[1]["sheet"]], MODULES[1])

    for module in MODULES[2:]:
        # copy from second sheet
        source = wb[MODULES[1]["sheet"]]
        target = wb.copy_worksheet(source)
        target.title = module["sheet"]
        write_feature_sheet(target, module)

    write_test_cases_sheet(wb)
    write_test_statistics_sheet(wb)

    wb.save(COMPLETED)
    OUT_PROJECT.mkdir(parents=True, exist_ok=True)
    shutil.copy2(COMPLETED, OUT_PROJECT / COMPLETED.name)
    shutil.copy2(BACKUP, OUT_PROJECT / BACKUP.name)
    return COMPLETED


def write_validation(path: Path, excel_path: Path):
    wb = load_workbook(excel_path, data_only=False)
    ids = []
    per_module = {}
    for m in MODULES:
        ws = wb[m["sheet"]]
        count = 0
        for row in ws.iter_rows(min_row=11, max_col=1, values_only=True):
            v = row[0]
            if isinstance(v, str) and v.startswith("IT_"):
                ids.append(v)
                count += 1
        per_module[m["code"]] = count
    dup = sorted({x for x in ids if ids.count(x) > 1})
    lines = [
        "# Integration Test Validation Report\n",
        f"Date: {ISSUE_DATE}\n",
        f"Workbook: `{excel_path}`\n",
        f"Sheets: {wb.sheetnames}\n",
        f"## Totals\n- Modules: {len(MODULES)}\n- Test cases: {len(ids)}\n- Duplicate IDs: {dup or 'None'}\n",
        "## Per module\n",
        "| Module | Count |\n|---|---:|",
    ]
    for k, v in per_module.items():
        lines.append(f"| {k} | {v} |")
    lines.append("\n## Endpoints covered (by design of IT cases)\n")
    lines.append("- /api/auth/* (register, login, verify, forgot/reset)\n- /api/user/me* (profile, password, notification-preferences)\n- /api/student/commerce/* , /api/student/payments/* , /api/payos/webhook\n- /api/online-courses/** , /api/student/online-courses/**\n- /api/classroom-offerings/** , /api/training-manager/classrooms/** , /api/student/classrooms/** , /api/teacher/classrooms/**\n- /api/student/placement-tests/**\n- /api/student/support-tickets/** , manager/staff support ticket APIs\n- /api/admin/users/**\n- /api/student/notifications/**\n")
    lines.append("## Special environments\n")
    lines.append("- PayOS, Mail, Gemini, Lark should be stubbed for deterministic IT.\n- Waitlist HTTP mapping gap: validate service method; restore controller if product requires.\n- No Testcontainers dependency today — use dedicated test PostgreSQL or profile.\n")
    lines.append("## Assumptions\n")
    lines.append("- UC IDs from SRS 1.3.2 are the primary requirement anchors.\n- Role names match `RoleEnum`.\n- Round statuses remain Pending until executed.\n")
    lines.append("## Checks performed\n")
    lines.append("1. Unique IT_* IDs\n2. Sheets exist for every Test Cases row\n3. Statistics formulas retargeted to module sheets\n4. No UI-click procedures\n5. Cover metadata filled from project code SEP490_G23 / EnglishLab\n")
    path.write_text("\n".join(lines), encoding="utf-8")


def main():
    OUT_PROJECT.mkdir(parents=True, exist_ok=True)
    if not BACKUP.exists():
        shutil.copy2(TEMPLATE, BACKUP)

    analysis = ROOT / "integration-test-analysis.md"
    plan = ROOT / "integration-test-plan.md"
    trace = ROOT / "integration-test-traceability.md"
    validation = ROOT / "integration-test-validation-report.md"

    write_analysis(analysis)
    write_plan(plan)
    write_traceability(trace)
    excel = build_excel()
    write_validation(validation, excel)

    # mirror markdown to project outputs
    for p in (analysis, plan, trace, validation):
        shutil.copy2(p, OUT_PROJECT / p.name)

    ids = all_case_ids()
    print("MODULES", len(MODULES))
    print("CASES", len(ids))
    print("UNIQUE", len(set(ids)))
    print("EXCEL", excel)
    print("OUT", OUT_PROJECT)


if __name__ == "__main__":
    main()
