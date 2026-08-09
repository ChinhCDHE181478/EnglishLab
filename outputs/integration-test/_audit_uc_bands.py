# -*- coding: utf-8 -*-
import shutil
from pathlib import Path

from openpyxl import load_workbook

from uc_modules import MODULES, iter_cases

srs_master = {
    "UC-01": "Register Account",
    "UC-02": "View public courses",
    "UC-03": "Login",
    "UC-04": "Reset password",
    "UC-05": "Manage profile",
    "UC-06": "View Notifications",
    "UC-07": "Submit Support Ticket",
    "UC-08": "Enroll in Course",
    "UC-09": "View Timetable",
    "UC-10": "Join Online Meeting",
    "UC-11": "Access Classroom Learning Materials",
    "UC-12": "Download Learning Materials",
    "UC-13": "Submit Homework",
    "UC-14": "View Academic Report",
    "UC-15": "Take Quiz",
    "UC-16": "Take Placement Exam",
    "UC-22": "View Teaching Schedule",
    "UC-23": "Manage Class Attendance",
    "UC-26": "Manage Homework",
    "UC-27": "Manage Quiz Practice Content",
    "UC-32": "Manage Syllabus",
    "UC-33": "Manage Online Courses",
    "UC-36": "Manage Classrooms",
    "UC-37": "Assign Teacher to Classroom",
    "UC-38": "Assign Learner to Classroom",
    "UC-40": "View operational report",
    "UC-41": "View revenue analytic of online course",
    "UC-42": "Manage User Accounts",
    "UC-43": "Manage System Notifications",
    "UC-44": "Resolve Support Tickets",
    "UC-45": "Wishlist Courses",
    "UC-46": "Add Courses to Cart",
    "UC-47": "Checkout",
    "UC-48": "Access Online Learning Materials",
}
srs_detail = {
    "UC-05a": "View profile",
    "UC-05b": "Update profile",
    "UC-23a": "View Class Attendance",
    "UC-23b": "Record Class Attendance",
    "UC-23c": "Update Class Attendance",
    "UC-26a": "Create Homework",
    "UC-26b": "View Homework",
    "UC-26c": "Update Homework",
    "UC-26d": "Delete Homework",
    "UC-27a": "Create Quiz Practice",
    "UC-27b": "View Quiz Practice",
    "UC-27c": "Update Quiz Practice",
    "UC-27d": "Delete Quiz Practice",
    "UC-32a": "Create Syllabus",
    "UC-32b": "View Syllabus",
    "UC-32c": "Update Syllabus",
    "UC-32d": "Delete Syllabus",
    "UC-33a": "Create Online Course",
    "UC-33b": "View Online Courses",
    "UC-33c": "Update Online Course",
    "UC-33d": "Deactivate Online Course",
    "UC-36a": "Create Classroom",
    "UC-36b": "View Classrooms",
    "UC-36c": "Update Classroom",
    "UC-36d": "Delete Classroom",
    "UC-42a": "Create User Account",
    "UC-42b": "View User Accounts",
    "UC-42c": "Update User Account",
    "UC-42d": "Lock/Unlock User Account",
    "UC-43a": "Create System Notification",
    "UC-43b": "View System Notifications",
    "UC-43c": "Update System Notification",
    "UC-43d": "Delete System Notification",
}
allowed = set(srs_master.values()) | set(srs_detail.values())

candidates = [
    Path(
        r"C:\Users\phong\Downloads\intergration test"
        r"\SEP490_G23_Report5.2_Integration Test_COMPLETED_UC_BANDS.xlsx"
    ),
    Path(
        r"C:\Users\phong\Downloads\intergration test"
        r"\SEP490_G23_Report5.2_Integration Test_COMPLETED.xlsx"
    ),
]
p = next(x for x in candidates if x.exists())
print("FILE", p)
wb = load_workbook(p, data_only=False)
bad = []
ok = []
for name in wb.sheetnames:
    if not name.startswith("IT_"):
        continue
    ws = wb[name]
    for r in range(11, min(ws.max_row or 11, 200) + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and not v.startswith("IT_"):
            (ok if v in allowed else bad).append((name, r, v))

print("OK bands", len(ok))
print("BAD bands", len(bad))
for x in bad:
    print(" ", x)

print("--- module bands ---")
for m in MODULES:
    bands = [g for k, g, _, c in iter_cases(m) if k == "GROUP"]
    print(m["sheet"], "|", m["function"], "|", bands)

src = Path(
    r"C:\Users\phong\Downloads\intergration test"
    r"\SEP490_G23_Report5.2_Integration Test_COMPLETED_UC_BANDS.xlsx"
)
dst = Path(
    r"C:\Users\phong\Downloads\intergration test"
    r"\SEP490_G23_Report5.2_Integration Test_COMPLETED.xlsx"
)
if src.exists():
    try:
        shutil.copy2(src, dst)
        print("OVERWROTE COMPLETED.xlsx")
    except Exception as e:
        print("CANNOT overwrite COMPLETED:", type(e).__name__, str(e))
