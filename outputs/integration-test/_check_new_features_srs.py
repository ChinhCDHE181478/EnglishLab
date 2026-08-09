# -*- coding: utf-8 -*-
from pathlib import Path

from openpyxl import load_workbook

from uc_modules import MODULES

SRS = {
    "UC-22": "View Teaching Schedule",
    "UC-23": "Manage Class Attendance",
    "UC-23a": "View Class Attendance",
    "UC-23b": "Record Class Attendance",
    "UC-26": "Manage Homework",
    "UC-26a": "Create Homework",
    "UC-13": "Submit Homework",
}
ALLOWED = set(SRS.values())

p = Path(
    r"C:\Users\phong\Downloads\intergration test"
    r"\SEP490_G23_Report5.2_Integration Test_COMPLETED.xlsx"
)
wb = load_workbook(p, data_only=True)

for code in ("SCHEDULE", "ATTEND", "MNGHW", "HOMEWORK"):
    m = next(x for x in MODULES if x["code"] == code)
    ws = wb[m["sheet"]]
    feat = ws["B2"].value
    bands = []
    for r in range(11, 30):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and not v.startswith("IT_"):
            bands.append(v)
    print(m["sheet"])
    print("  Feature Excel :", repr(feat), "| match SRS?", feat in ALLOWED)
    print("  uc_modules    :", repr(m["function"]))
    print("  ucs           :", m["ucs"])
    print("  cyan bands    :", bands)
    for b in bands:
        print("   ", repr(b), "| match SRS?", b in ALLOWED)
    print()
