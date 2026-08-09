# -*- coding: utf-8 -*-
"""Remove 'MODULE N - ' prefix from group-band cells in column A (Test Case ID)."""
from __future__ import annotations

import re
import shutil
from pathlib import Path

from openpyxl import load_workbook

EXCEL = Path(
    r"C:\Users\phong\Downloads\intergration test"
    r"\SEP490_G23_Report5.2_Integration Test_COMPLETED.xlsx"
)
PROJ = Path(r"D:\EngLishLab\EnglishLab\outputs\integration-test")
BACKUP = EXCEL.with_name(EXCEL.stem + "_BEFORE_STRIP_MODULE.xlsx")

# MODULE 1 - Title  /  MODULE 12 - Title
PAT = re.compile(r"^MODULE\s+\d+\s*-\s*(.+)$", re.IGNORECASE)


def main() -> None:
    if not EXCEL.exists():
        raise SystemExit(f"Missing {EXCEL}")
    shutil.copy2(EXCEL, BACKUP)
    wb = load_workbook(EXCEL)
    changed = 0
    for name in wb.sheetnames:
        if not name.startswith("IT_"):
            continue
        ws = wb[name]
        for row in range(11, min(ws.max_row or 11, 500) + 1):
            val = ws.cell(row, 1).value
            if not isinstance(val, str):
                continue
            m = PAT.match(val.strip())
            if not m:
                continue
            new = m.group(1).strip()
            ws.cell(row, 1).value = new
            changed += 1
            print(f"{name}!A{row}: {val!r} -> {new!r}")

    try:
        wb.save(EXCEL)
        out = EXCEL
    except PermissionError:
        out = EXCEL.with_name(EXCEL.stem + "_NO_MODULE.xlsx")
        wb.save(out)
        print("LOCKED original ->", out)

    shutil.copy2(out, PROJ / out.name)
    print("changed", changed, "BACKUP", BACKUP, "WROTE", out)


if __name__ == "__main__":
    main()
