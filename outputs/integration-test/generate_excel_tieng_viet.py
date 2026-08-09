# -*- coding: utf-8 -*-
"""
Tao ban Excel Integration Test tieng Viet (de doc) tu COMPLETED.xlsx.

- Dich bang Google Translate (deep-translator), bao ve API / HTTP / ten class Java.
- Dich nhan bang map co dinh (Feature, header, trang thai).
- Giu nguyen: Test Case ID, ngay/tester, cong thuc, hyperlink.
"""
from __future__ import annotations

import hashlib
import json
import re
import shutil
import time
from pathlib import Path

from deep_translator import GoogleTranslator
from openpyxl import load_workbook

SRC = Path(
    r"C:\Users\phong\Downloads\intergration test"
    r"\SEP490_G23_Report5.2_Integration Test_COMPLETED.xlsx"
)
OUT_PROJ = Path(
    r"D:\EngLishLab\EnglishLab\outputs\integration-test"
    r"\SEP490_G23_Report5.2_Integration Test_COMPLETED_TIENG_VIET.xlsx"
)
OUT_DL = Path(
    r"C:\Users\phong\Downloads\intergration test"
    r"\SEP490_G23_Report5.2_Integration Test_COMPLETED_TIENG_VIET.xlsx"
)
CACHE_PATH = Path(
    r"D:\EngLishLab\EnglishLab\outputs\integration-test\_vi_translate_cache.json"
)

STATUS_MAP = {
    "Passed": "Đạt",
    "Failed": "Không đạt",
    "Pending": "Chờ",
    "N/A": "Không áp dụng",
    "NA": "Không áp dụng",
}

FEATURE_MAP = {
    "View public courses": "Xem khóa học công khai",
    "Access Online Learning Materials": "Truy cập tài liệu học online",
    "Assign Learner to Classroom": "Xếp học viên vào lớp",
    "Enroll in Course": "Đăng ký khóa học",
    "Wishlist Courses": "Danh sách yêu thích khóa học",
    "Add Courses to Cart": "Thêm khóa học vào giỏ hàng",
    "Checkout": "Thanh toán (Checkout)",
    "Register Account": "Đăng ký tài khoản",
    "Manage profile": "Quản lý hồ sơ",
    "Manage Classrooms": "Quản lý lớp học",
    "Assign Teacher to Classroom": "Phân công giáo viên vào lớp",
    "View Teaching Schedule": "Xem lịch giảng dạy",
    "Manage Class Attendance": "Quản lý điểm danh lớp",
    "Manage Homework": "Quản lý bài tập (giáo viên)",
    "View Timetable": "Xem thời khóa biểu",
    "Access Classroom Learning Materials": "Truy cập tài liệu học của lớp",
    "Submit Homework": "Nộp bài tập",
    "Take Quiz": "Làm bài quiz",
    "Take Placement Exam": "Làm bài kiểm tra xếp lớp",
    "Manage Online Courses": "Quản lý khóa học online",
    "Manage Syllabus": "Quản lý giáo trình (Syllabus)",
    "View Notifications": "Xem thông báo",
    "Submit Support Ticket": "Gửi phiếu hỗ trợ",
    "Manage User Accounts": "Quản lý tài khoản người dùng",
    "Manage System Notifications": "Quản lý thông báo hệ thống",
    "Join Online Meeting": "Tham gia buổi học online (Meet)",
    "View operational report": "Xem báo cáo vận hành",
    "Login": "Đăng nhập",
    "Reset password": "Đặt lại mật khẩu",
    "Create Classroom": "Tạo lớp học",
    "Update Classroom": "Cập nhật lớp học",
    "View Classrooms": "Xem danh sách lớp học",
    "Create Homework": "Tạo bài tập",
    "Record Class Attendance": "Ghi nhận điểm danh",
    "View Class Attendance": "Xem điểm danh",
    "Create Online Course": "Tạo khóa học online",
    "Update Online Course": "Cập nhật khóa học online",
    "View Online Courses": "Xem khóa học online",
    "Create Syllabus": "Tạo giáo trình",
    "Update Syllabus": "Cập nhật giáo trình",
    "Delete Syllabus": "Xóa giáo trình",
    "View Syllabus": "Xem giáo trình",
    "Create System Notification": "Tạo thông báo hệ thống",
    "Update System Notification": "Cập nhật thông báo hệ thống",
    "Delete System Notification": "Xóa thông báo hệ thống",
    "View System Notifications": "Xem thông báo hệ thống",
    "Lock/Unlock User Account": "Khóa / mở khóa tài khoản",
    "View User Accounts": "Xem danh sách tài khoản",
    "Resolve Support Tickets": "Xử lý phiếu hỗ trợ",
    "Create Quiz Practice": "Tạo bài quiz luyện tập",
    "View revenue analytic of online course": "Xem phân tích doanh thu khóa online",
}

LABEL_MAP = {
    "TEST REPORT DOCUMENT": "TÀI LIỆU BÁO CÁO KIỂM THỬ",
    "TEST CASE LIST": "DANH SÁCH TEST CASE",
    "TEST STATISTICS": "THỐNG KÊ KIỂM THỬ",
    "Project Name": "Tên dự án",
    "Project Code": "Mã dự án",
    "Document Code": "Mã tài liệu",
    "Creator": "Người tạo",
    "Issue Date": "Ngày ban hành",
    "Version": "Phiên bản",
    "Reviewer/Approver": "Người review / phê duyệt",
    "Notes": "Ghi chú",
    "Record of change": "Lịch sử thay đổi",
    "Effective Date": "Ngày hiệu lực",
    "Change Item": "Hạng mục thay đổi",
    "Change description": "Mô tả thay đổi",
    "Reference": "Tham chiếu",
    "Feature": "Chức năng (Feature)",
    "Test requirement": "Yêu cầu kiểm thử",
    "Number of TCs": "Số lượng TC",
    "Testing Round": "Vòng kiểm thử",
    "Round 1": "Vòng 1",
    "Round 2": "Vòng 2",
    "Round 3": "Vòng 3",
    "Test Case ID": "Mã Test Case",
    "Test Case Description": "Mô tả Test Case",
    "Test Case Procedure": "Các bước thực hiện",
    "Expected Results": "Kết quả mong đợi",
    "Pre-conditions": "Tiền điều kiện",
    "Test date": "Ngày test",
    "Tester": "Người test",
    "Note": "Ghi chú",
    "Test Environment Setup Description": "Mô tả môi trường kiểm thử",
    "Function Name": "Tên chức năng",
    "No": "STT",
    "Module code": "Mã module",
    "Passed": "Đạt",
    "Failed": "Không đạt",
    "Pending": "Chờ",
    "N/A": "Không áp dụng",
}

translator = GoogleTranslator(source="en", target="vi")
_cache: dict[str, str] = {}


def load_cache() -> None:
    global _cache
    if CACHE_PATH.exists():
        _cache = json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    else:
        _cache = {}


def save_cache() -> None:
    CACHE_PATH.write_text(json.dumps(_cache, ensure_ascii=False, indent=0), encoding="utf-8")


def _cache_key(text: str) -> str:
    return hashlib.sha1(text.encode("utf-8")).hexdigest()


# Single-pass alternation (order = priority). Placeholders use [[#n]] so they
# are never re-matched by technical patterns.
_PROTECT_RE = re.compile(
    r"("
    r"/api/[A-Za-z0-9_\-/{}.?=*&]+"
    r"|HTTP\s*\d{3}(?:\s*OK|\s*Bad Request|\s*Forbidden|\s*Not Found)?"
    r"|\bIT_[A-Z0-9_]+\b"
    r"|\b(?:PUBLISHED|DRAFT|SUBMITTED|WAITING_FOR_CLASS|CLASS_ASSIGNED|"
    r"LEARNER|STAFF|ADMIN|TEACHER|CM|JWT|OTP|N/A|MockMvc|PayOS|"
    r"Google Meet|Content-Type|application/json)\b"
    r"|\b[A-Za-z][A-Za-z0-9]*(?:\.[A-Za-z][A-Za-z0-9]*)+\(\)"
    r"|\b[A-Z][A-Za-z0-9]+(?:Controller|Service|ServiceImpl|Repository|Request|Response|FilterChain)\b"
    r"|\b[A-Za-z][A-Za-z0-9]*\(\)"
    r"|\{[A-Za-z0-9_]+\}"
    r"|__(?!TK)[A-Za-z0-9_]+__"
    r"|@[A-Za-z][A-Za-z0-9]*"
    r"|\b(?:sendInApp|sendEmail|accessToken|fullName|classroomId|lessonId|"
    r"courseId|requestId|teacherId|totalElements|email_verified)=?(?:true|false)?\b"
    r")"
)


def protect_technical(text: str) -> tuple[str, dict[str, str]]:
    mapping: dict[str, str] = {}
    idx = 0

    def repl(m: re.Match) -> str:
        nonlocal idx
        token = m.group(0)
        key = f"[[#{idx}]]"
        idx += 1
        mapping[key] = token
        return key

    out = _PROTECT_RE.sub(repl, text)
    return out, mapping


def restore_technical(text: str, mapping: dict[str, str]) -> str:
    out = text
    for key, val in sorted(mapping.items(), key=lambda x: -len(x[0])):
        if key in out:
            out = out.replace(key, val)
            continue
        n = re.search(r"(\d+)", key).group(1)
        out, count = re.subn(
            rf"\[\[\s*#\s*{n}\s*\]\]", val, out, count=1
        )
        if count:
            continue
        out, _ = re.subn(rf"#\s*{n}", val, out, count=1)
    return out


def google_translate(text: str) -> str:
    key = _cache_key(text)
    if key in _cache:
        return _cache[key]

    protected, mapping = protect_technical(text)
    if len(protected) <= 4200:
        parts = [protected]
    else:
        parts = []
        buf = []
        size = 0
        for line in protected.splitlines(keepends=True):
            if size + len(line) > 3500 and buf:
                parts.append("".join(buf))
                buf = [line]
                size = len(line)
            else:
                buf.append(line)
                size += len(line)
        if buf:
            parts.append("".join(buf))

    chunks = []
    for part in parts:
        attempt = 0
        while True:
            try:
                vi = translator.translate(part)
                chunks.append(vi)
                break
            except Exception as e:
                attempt += 1
                if attempt >= 5:
                    raise RuntimeError(f"Translate failed: {e}") from e
                time.sleep(1.2 * attempt)
        time.sleep(0.08)

    joined = "".join(chunks)
    restored = restore_technical(joined, mapping)
    restored = polish_vi(restored)
    _cache[key] = restored
    return restored


def polish_vi(text: str) -> str:
    """Sua mot so loi dich thuong gap de de doc hon."""
    reps = [
        (r"tiêu đề Ủy quyền", "header Authorization"),
        (r"tiêu đề ủy quyền", "header Authorization"),
        (r"header ủy quyền", "header Authorization"),
        (r"Ủy quyền Bearer", "Authorization Bearer"),
        (r"ủy quyền Bearer", "Authorization Bearer"),
        (r"mã thông báo Người mang ủy quyền", "token Authorization Bearer"),
        (r"Người mang ủy quyền", "Authorization Bearer"),
        (r"nhận được yêu cầu và ủy quyền cho", "nhận request và gọi tiếp"),
        (r"nhận yêu cầu và ủy quyền cho", "nhận request và gọi tiếp"),
        (r"ủy quyền cho", "gọi tiếp"),
        (r"đại biểu cho", "gọi tiếp"),
        (r"Đại biểu cho", "Gọi tiếp"),
        (r"Bộ điều khiển-Dịch vụ-Kho lưu trữ", "Controller–Service–Repository"),
        (r"Bộ điều khiển", "Controller"),
        (r"Trạng thái HTTP là", "HTTP status là"),
        (r"trạng thái HTTP là", "HTTP status là"),
        (r"^CNTT để ", "IT cho "),
        (r"^CNTT dành cho ", "IT cho "),
        (r"\bCNTT\b", "IT"),
        (r"thông qua Controller", "qua Controller"),
        (r"văn bản thuần túy", "plain text"),
        (r"được băm", "được hash"),
        (r"dạng băm", "dạng hash"),
        (r"tải trọng", "payload"),
        # without Authorization = public / no auth required (not "không được phép")
        (r"qua MockMvc mà không được phép", "qua MockMvc (không cần xác thực)"),
        (r"qua MockMvc mà không có header Authorization", "qua MockMvc (không cần xác thực)"),
        (r"mà không có sự cho phép", " (không cần xác thực)"),
        (r"không có sự cho phép", "không cần xác thực"),
        (r"không có header Authorization", "không cần xác thực"),
        (r"without Authorization header", "không cần xác thực"),
        (r"without Authorization", "không cần xác thực"),
    ]
    out = text
    for pat, repl in reps:
        out = re.sub(pat, repl, out)
    # plain string cleanup (do NOT use regex '.' here)
    out = out.replace(" .", ".").replace(" ,", ",")
    out = re.sub(r"[ \t]{2,}", " ", out)
    return out


def translate_cell(value, *, force_machine: bool = False):
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    if value.startswith("="):
        return value

    s = value.strip()
    if s in STATUS_MAP and "\n" not in value:
        return STATUS_MAP[s]
    if "\n" not in value and s in LABEL_MAP:
        return LABEL_MAP[s]
    if "\n" not in value and s in FEATURE_MAP:
        return FEATURE_MAP[s]
    if s in ("None.", "None"):
        return "Không." if s.endswith(".") else "Không"

    if not re.search(r"[A-Za-z]", value):
        return value

    # short exact labels already handled; machine-translate narrative
    if not force_machine and "\n" not in value and len(s) < 40 and s in LABEL_MAP:
        return LABEL_MAP[s]

    return google_translate(value)


def translate_formula_status_refs(formula: str) -> str:
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula
    f = formula
    for en, vi in STATUS_MAP.items():
        f = f.replace(f'"{en}"', f'"{vi}"')
    return f


def process_module_sheet(ws) -> int:
    changed = 0
    max_r = ws.max_row or 1
    max_c = min(ws.max_column or 15, 15)

    for r in range(1, 11):
        for c in range(1, max_c + 1):
            cell = ws.cell(r, c)
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                nv = translate_formula_status_refs(v)
                if nv != v:
                    cell.value = nv
                    changed += 1
                continue
            if v is None or v == "":
                continue
            force = r == 3 and c == 2
            nv = translate_cell(v, force_machine=force)
            if nv != v and nv is not None:
                cell.value = nv
                changed += 1

    for r in range(11, max_r + 1):
        a = ws.cell(r, 1).value
        if isinstance(a, str) and a.strip() and not a.strip().startswith("IT_"):
            nv = translate_cell(a)
            if nv != a:
                ws.cell(r, 1).value = nv
                changed += 1
            continue
        if not (isinstance(a, str) and a.startswith("IT_")):
            continue

        for c in (2, 3, 4, 5, 15):
            cell = ws.cell(r, c)
            if cell.value is None or cell.value == "":
                continue
            nv = translate_cell(cell.value, force_machine=True)
            if nv != cell.value and nv is not None:
                cell.value = nv
                changed += 1

        for c in (6, 9, 12):
            cell = ws.cell(r, c)
            if isinstance(cell.value, str) and cell.value.strip() in STATUS_MAP:
                cell.value = STATUS_MAP[cell.value.strip()]
                changed += 1
    return changed


def process_generic(ws, max_row=200, max_col=15) -> int:
    changed = 0
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, max_row), max_col=max_col):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                nv = translate_formula_status_refs(v)
                if nv != v:
                    cell.value = nv
                    changed += 1
                continue
            if v is None or v == "":
                continue
            force = isinstance(v, str) and ("\n" in v or len(v) > 40)
            nv = translate_cell(v, force_machine=force)
            if nv != v and nv is not None:
                cell.value = nv
                changed += 1
    return changed


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"Missing source: {SRC}")

    load_cache()
    OUT_PROJ.parent.mkdir(parents=True, exist_ok=True)
    OUT_DL.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SRC, OUT_PROJ)

    wb = load_workbook(OUT_PROJ)
    total = 0

    if "Cover" in wb.sheetnames:
        n = process_generic(wb["Cover"], max_row=40, max_col=10)
        print(f"Cover: {n}", flush=True)
        total += n
        cover = wb["Cover"]
        cover["A13"] = "Ngôn ngữ"
        cover["B13"] = (
            "Tiếng Việt — dịch dễ đọc từ COMPLETED.xlsx; "
            "giữ nguyên Test Case ID, đường API, mã HTTP, ngày test / tester."
        )

    if "Test Cases" in wb.sheetnames:
        n = process_generic(wb["Test Cases"], max_row=80, max_col=10)
        print(f"Test Cases: {n}", flush=True)
        total += n

    if "Test Statistics" in wb.sheetnames:
        n = process_generic(wb["Test Statistics"], max_row=120, max_col=15)
        print(f"Test Statistics: {n}", flush=True)
        total += n

    for name in wb.sheetnames:
        if not name.startswith("IT_"):
            continue
        n = process_module_sheet(wb[name])
        print(f"{name}: {n}", flush=True)
        total += n
        save_cache()

    save_cache()
    # Save once at the end (repeated saves break embedded images in openpyxl).
    wb.save(OUT_PROJ)
    shutil.copy2(OUT_PROJ, OUT_DL)
    print(f"DONE cells_changed~={total}", flush=True)
    print(f"OUT: {OUT_PROJ}", flush=True)
    print(f"OUT: {OUT_DL}", flush=True)


if __name__ == "__main__":
    main()
